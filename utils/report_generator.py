"""
utils/report_generator.py

Three report formats — PDF, Excel (.xlsx), HTML — all sharing one design system.

Design tokens
─────────────
Primary navy  #1E3A8A  — header bars, table headers
Accent blue   #3B82F6  — section underlines, borders
Red title     #DC2626  — "RESULT ANALYSIS REPORT" subtitle
Dark text     #0F172A  — body text
Light surface #F1F5F9  — alternate rows, overview bg
Mid border    #CBD5E1  — table grid

Grade colour language (same meaning in every format)
─────────────────────────────────────────────────────
S / A+       deep green   → outstanding
A  / B+      light green  → very good
B            yellow       → good
C+ / C       amber/orange → average
D            salmon       → below average
P            sky blue     → pass (practical)
F / FE       solid red    → fail
"""

import os
import base64
from datetime import datetime

# ── ReportLab ────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable, PageBreak,
    Image, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as rl
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── openpyxl ─────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage


# ═══════════════════════════════════════════════════════════════
# SHARED DESIGN TOKENS
# ═══════════════════════════════════════════════════════════════

# Primary palette
_NAVY  = "#1E3A8A"
_BLUE  = "#3B82F6"
_RED   = "#DC2626"
_DARK  = "#0F172A"
_LGRAY = "#F1F5F9"
_MGRAY = "#CBD5E1"
_WHITE = "#FFFFFF"

# Grade colour palette — bg / text
GRADE_COLORS = {
    "S":      (_WHITE, "#166534", "#166534"),  # (xlsx_bg, xlsx_fg, html_bg) — split below
    "A+":     ("#15803D", _WHITE, "#15803D"),
    "A":      ("#86EFAC", "#14532D", "#86EFAC"),
    "B+":     ("#FEF08A", "#713F12", "#FEF08A"),
    "B":      ("#FDE047", "#713F12", "#FDE047"),
    "C+":     ("#FED7AA", "#7C2D12", "#FED7AA"),
    "C":      ("#FDBA74", "#7C2D12", "#FDBA74"),
    "D":      ("#FCA5A5", "#7F1D1D", "#FCA5A5"),
    "P":      ("#BAE6FD", "#0C4A6E", "#BAE6FD"),
    "F":      (_RED, _WHITE, _RED),
    "FE":     (_RED, _WHITE, _RED),
    "ABSENT": ("#94A3B8", "#0F172A", "#94A3B8"),
}

# Consistent grade → (bg_hex, fg_hex)
def _grade_colors(grade: str):
    """Return (bg_hex, fg_hex) for a grade string."""
    g = str(grade).strip().upper()
    entry = GRADE_COLORS.get(g)
    if entry:
        return entry[0], entry[1]
    return None, None

FAIL_GRADES = {"F", "FE", "ABSENT"}


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def safe_float(val) -> float:
    try:
        v = str(val).strip()
        return 0.0 if v in ("", "nan", "None", "-") else float(v)
    except (ValueError, TypeError):
        return 0.0


def _c(hex_str: str):
    """ReportLab HexColor shorthand."""
    return rl.HexColor(hex_str)


def _qr_b64(qr_path: str) -> str:
    """Encode QR PNG as data-URI for self-contained HTML."""
    if not qr_path or not os.path.exists(qr_path):
        return ""
    with open(qr_path, "rb") as f:
        return "data:image/png;base64," + base64.b64encode(f.read()).decode()


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_str: str):
    return PatternFill(start_color=hex_str.lstrip("#"),
                       end_color=hex_str.lstrip("#"), fill_type="solid")


# ═══════════════════════════════════════════════════════════════
# PDF REPORT
# ═══════════════════════════════════════════════════════════════

def generate_pdf_report(
    df, analysis_df, output_path, report_id, qr_path,
    department="", semester="", batch=""
):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df = df.fillna("")

    # ── Page geometry ─────────────────────────────────────────
    LM, RM = 0.55 * inch, 0.55 * inch
    TM, BM = 0.45 * inch, 0.55 * inch
    AW = A4[0] - LM - RM

    # ── Colour aliases ────────────────────────────────────────
    NAVY  = _c(_NAVY)
    BLUE  = _c(_BLUE)
    RED   = _c(_RED)
    DARK  = _c(_DARK)
    LGRAY = _c(_LGRAY)
    MGRAY = _c(_MGRAY)
    GOLD  = _c("#D97706")

    # ── Footer callback ───────────────────────────────────────
    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(_c("#64748B"))
        canvas.drawString(LM, BM * 0.42,
            "APJ Abdul Kalam Technological University  ·  "
            "Carmel College of Engineering and Technology")
        canvas.drawRightString(A4[0] - RM, BM * 0.42, f"Page {doc.page}")
        # thin top line on every page after first
        if doc.page > 1:
            canvas.setStrokeColor(MGRAY)
            canvas.setLineWidth(0.4)
            canvas.line(LM, A4[1] - TM * 0.6, A4[0] - RM, A4[1] - TM * 0.6)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=LM, rightMargin=RM,
        topMargin=TM, bottomMargin=BM,
    )

    # ── Style definitions ─────────────────────────────────────
    S = getSampleStyleSheet()

    def _ps(name, parent=None, **kw):
        return ParagraphStyle(name, parent=parent or S["Normal"], **kw)

    college_s = _ps("College", fontName="Helvetica-Bold", fontSize=13,
                    textColor=rl.white, alignment=TA_CENTER, leading=18,
                    spaceAfter=0, spaceBefore=0)
    red_s = _ps("RedTitle", fontName="Helvetica-Bold", fontSize=12,
                textColor=RED, alignment=TA_CENTER, leading=16,
                spaceAfter=0, spaceBefore=0)
    sect_s = _ps("Sect", fontName="Helvetica-Bold", fontSize=10.5,
                 textColor=DARK, leading=14, spaceBefore=0, spaceAfter=4)
    meta_key = _ps("MKey", fontName="Helvetica-Bold", fontSize=8.5,
                   textColor=DARK, alignment=TA_RIGHT, leading=14)
    meta_sep = _ps("MSep", fontName="Helvetica", fontSize=8.5,
                   textColor=DARK, alignment=TA_CENTER, leading=14)
    meta_val = _ps("MVal", fontName="Helvetica", fontSize=8.5,
                   textColor=DARK, alignment=TA_LEFT, leading=14)
    th_s = _ps("TH", fontName="Helvetica-Bold", fontSize=7.5,
               textColor=rl.white, alignment=TA_CENTER, leading=10)
    td_s = _ps("TD", fontName="Helvetica", fontSize=7.5,
               alignment=TA_CENTER, leading=10)
    td_l = _ps("TDL", fontName="Helvetica", fontSize=7.5,
               alignment=TA_LEFT, leading=10)
    td_b = _ps("TDB", fontName="Helvetica-Bold", fontSize=7.5,
               alignment=TA_CENTER, leading=10)
    ov_key = _ps("OVK", fontName="Helvetica-Bold", fontSize=9.5,
                 textColor=DARK, alignment=TA_LEFT, leading=13)
    ov_val = _ps("OVV", fontName="Helvetica-Bold", fontSize=9.5,
                 textColor=DARK, alignment=TA_CENTER, leading=13)
    top_rank = _ps("Rank", fontName="Helvetica-Bold", fontSize=9,
                   textColor=GOLD, alignment=TA_CENTER, leading=13)
    ver_head = _ps("VH", fontName="Helvetica-Bold", fontSize=11,
                   textColor=_c("#1D4ED8"), leading=16, spaceAfter=4)
    ver_body = _ps("VB", fontName="Helvetica", fontSize=8.5,
                   textColor=DARK, leading=13)
    sig_lbl = _ps("SL", fontName="Helvetica-Bold", fontSize=9,
                  textColor=DARK, alignment=TA_CENTER, leading=13)

    def blue_rule(thick=1.2):
        return HRFlowable(width="100%", thickness=thick,
                          color=BLUE, spaceBefore=2, spaceAfter=4)

    def section(title):
        return [Spacer(1, 10), Paragraph(title, sect_s), blue_rule(0.8)]

    # ── Pre-calculate ─────────────────────────────────────────
    subj_cols = [c for c in df.columns if c not in ("Register No","SGPA","Arrear")]
    sgpas     = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa  = round(sum(sgpas) / len(sgpas), 2) if sgpas else 0
    max_sgpa  = round(max(sgpas), 2) if sgpas else 0
    min_sgpa  = round(min(sgpas), 2) if sgpas else 0
    failures  = sum(1 for s in sgpas if s == 0)
    date_str  = datetime.now().strftime("%d %B %Y")
    sem_lbl   = f"S{semester}" if semester else ""

    els = []  # elements list

    # ══════════════════════════════════
    # § 1  HEADER
    # ══════════════════════════════════

    # Navy header bar (college name)
    hdr_bar = Table(
        [[Paragraph("CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY", college_s)]],
        colWidths=[AW],
    )
    hdr_bar.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), NAVY),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ("RIGHTPADDING",  (0,0), (-1,-1), 8),
    ]))
    els.append(hdr_bar)

    # Red subtitle bar
    sub_bar = Table(
        [[Paragraph("RESULT ANALYSIS REPORT", red_s)]],
        colWidths=[AW],
    )
    sub_bar.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), _c("#EFF6FF")),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LINEBELOW",     (0,0), (-1,-1), 2, BLUE),
        ("LINEABOVE",     (0,0), (-1,-1), 0.5, MGRAY),
    ]))
    els.append(sub_bar)
    els.append(Spacer(1, 6))

    # Meta (key : value table) + QR
    LKEY, LSEP = 62, 12
    LVAL = AW * 0.65 - LKEY - LSEP

    def _mrow(key, val, val_style=None):
        return [Paragraph(key, meta_key),
                Paragraph(":", meta_sep),
                Paragraph(val, val_style or meta_val)]

    # Report ID gets a smaller font so long dept names never wrap mid-word
    meta_id_val = ParagraphStyle(
        "MIDVal", parent=meta_val,
        fontName="Helvetica", fontSize=7.5, leading=11,
    )

    meta_rows = [
        _mrow("Report ID",  report_id, meta_id_val),
        _mrow("Date",       date_str),
        _mrow("Department", department),
        _mrow("Semester",   sem_lbl),
    ]
    if batch:
        meta_rows.append(_mrow("Batch", batch))

    meta_tbl = Table(meta_rows, colWidths=[LKEY, LSEP, LVAL])
    meta_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0), (-1,-1), 0),
        ("RIGHTPADDING",  (0,0), (-1,-1), 2),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,0),  5),
    ]))

    if qr_path and os.path.exists(qr_path):
        qr_tbl = Table(
            [[Image(qr_path, width=68, height=68)],
             [Paragraph("<para align='center'><font size='6.5' color='#64748B'>"
                        "Scan to verify<br/>authenticity</font></para>", S["Normal"])]],
            colWidths=[AW * 0.32],
        )
        qr_tbl.setStyle(TableStyle([
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ]))
    else:
        qr_tbl = Paragraph("", S["Normal"])

    info_row = Table([[meta_tbl, qr_tbl]], colWidths=[AW * 0.68, AW * 0.32])
    info_row.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0), (-1,-1), 0),
        ("RIGHTPADDING",  (0,0), (-1,-1), 0),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    els.append(info_row)
    els.append(HRFlowable(width="100%", thickness=2, color=BLUE,
                          spaceBefore=0, spaceAfter=14))

    # ══════════════════════════════════
    # § 2  RESULT OVERVIEW
    # ══════════════════════════════════

    els += section("RESULT OVERVIEW")

    ov_data = [
        [Paragraph("Total Students",    ov_key), Paragraph(str(len(df)), ov_val)],
        [Paragraph("Total Subjects",    ov_key), Paragraph(str(len(subj_cols)), ov_val)],
        [Paragraph("Average SGPA",      ov_key), Paragraph(str(avg_sgpa), ov_val)],
        [Paragraph("Highest SGPA",      ov_key), Paragraph(str(max_sgpa), ov_val)],
        [Paragraph("Lowest SGPA",       ov_key), Paragraph(str(min_sgpa), ov_val)],
        [Paragraph("Failures (SGPA 0)", ov_key),
         Paragraph(str(failures), ParagraphStyle(
             "F0", parent=ov_val, fontName="Helvetica-Bold", fontSize=9.5,
             textColor=RED, alignment=TA_CENTER, leading=13))],
    ]
    ov_tbl = Table(ov_data, colWidths=[AW * 0.46, AW * 0.2], hAlign="LEFT")
    ov_tbl.setStyle(TableStyle([
        ("GRID",          (0,0), (-1,-1), 0.5, MGRAY),
        ("ROWBACKGROUNDS",(0,0), (-1,-1), [rl.white, LGRAY]),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (0,-1),  8),
        ("LEFTPADDING",   (1,0), (1,-1),  4),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("LINEBELOW",     (0,-1),(-1,-1), 1.5, BLUE),
    ]))
    els.append(ov_tbl)
    els.append(Spacer(1, 6))

    # ══════════════════════════════════
    # § 3  STUDENTS RESULT
    # ══════════════════════════════════

    els += section("STUDENTS RESULT")

    # Column widths: reg_no | subjects... | SGPA | Arrear
    REG_W  = 60
    SGPA_W = 30
    ARR_W  = 28
    SUBJ_W = (AW - REG_W - SGPA_W - ARR_W) / max(1, len(subj_cols))
    col_ws = [REG_W] + [SUBJ_W] * len(subj_cols) + [SGPA_W, ARR_W]

    hdr_row = [Paragraph(h, th_s) for h in df.columns]
    body_rows = []
    rs = []  # per-cell style overrides

    for i, (_, row) in enumerate(df.iterrows()):
        sgpa      = safe_float(row.get("SGPA", 0))
        ri        = i + 1  # row index in full table (0 = header)
        reg_color = RED if sgpa == 0 else DARK
        row_bg    = (_c("#FFEBEB") if sgpa == 0
                     else (_c("#FFFBEB") if 0 < sgpa < 5.0
                     else rl.white))
        rs.append(("BACKGROUND", (0, ri), (-1, ri), row_bg))

        cells = []
        for j, col in enumerate(df.columns):
            val = str(row[col]).strip()
            if col == "Register No":
                st2 = ParagraphStyle(f"R{i}", parent=td_l,
                                     fontName="Helvetica-Bold",
                                     textColor=RED if sgpa == 0 else DARK)
                cells.append(Paragraph(val, st2))
            elif col in subj_cols:
                bg_hex, fg_hex = _grade_colors(val.upper())
                if bg_hex:
                    rs.append(("BACKGROUND", (j, ri), (j, ri), _c(bg_hex)))
                    st2 = ParagraphStyle(f"G{i}{j}", parent=td_b,
                                         textColor=_c(fg_hex))
                    cells.append(Paragraph(val, st2))
                else:
                    cells.append(Paragraph(val, td_s))
            elif col == "SGPA":
                st2 = ParagraphStyle(f"S{i}", parent=td_b,
                                     textColor=RED if sgpa == 0
                                     else (_c("#16A34A") if sgpa >= 7.5
                                     else DARK))
                cells.append(Paragraph(val, st2))
            else:
                cells.append(Paragraph(val, td_s))
        body_rows.append(cells)

    base = [
        ("BACKGROUND",    (0,0), (-1,0),  NAVY),
        ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 2),
        ("RIGHTPADDING",  (0,0), (-1,-1), 2),
        ("REPEATROWS",    (0,0), (0,0)),
    ]
    stud_tbl = Table([hdr_row] + body_rows, colWidths=col_ws,
                     hAlign="LEFT", repeatRows=1)
    stud_tbl.setStyle(TableStyle(base + rs))
    els.append(stud_tbl)
    els.append(Spacer(1, 8))

    # ══════════════════════════════════
    # § 4  TOP PERFORMERS  (no page split)
    # ══════════════════════════════════

    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)

    medals = ["🥇", "🥈", "🥉"]
    top_hdr = [Paragraph(h, th_s) for h in ["Rank", "Register No", "SGPA"]]
    top_rows = []
    for rank, (_, row) in enumerate(top_df.iterrows(), 1):
        top_rows.append([
            Paragraph(str(rank), top_rank),
            Paragraph(row["Register No"], td_l),
            Paragraph(str(row["SGPA"]), td_b),
        ])

    top_tbl = Table([top_hdr] + top_rows,
                    colWidths=[38, AW * 0.5, 48], hAlign="LEFT")
    top_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  _c("#D97706")),
        ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [rl.white, _c("#FFFBEB")]),
        ("LINEBELOW",     (0,-1),(-1,-1), 1.5, _c("#D97706")),
    ]))

    els.append(KeepTogether(section("TOP PERFORMERS") + [top_tbl, Spacer(1, 8)]))

    # ══════════════════════════════════
    # § 5  SUBJECT PERFORMANCE ANALYSIS
    #       — flows naturally after Top Performers (no forced page break)
    # ══════════════════════════════════

    els += section("SUBJECT PERFORMANCE ANALYSIS")

    ana_cols = list(analysis_df.columns)
    SUBJ_NW = 64
    OTHER_W = (AW - SUBJ_NW) / max(1, len(ana_cols) - 1)
    ana_ws  = [SUBJ_NW] + [OTHER_W] * (len(ana_cols) - 1)

    ana_hdr = [Paragraph(h, th_s) for h in ana_cols]
    ana_rows = []
    for _, row in analysis_df.iterrows():
        cells = []
        for col in ana_cols:
            v = str(row[col])
            if col == "Subject":
                cells.append(Paragraph(v, td_l))
            elif "%" in v:
                try:
                    pct = float(v.replace("%", ""))
                    fg = (RED if pct < 60
                          else (_c("#D97706") if pct < 80
                          else _c("#16A34A")))
                    st2 = ParagraphStyle(f"PP{col}", parent=td_b, textColor=fg)
                    cells.append(Paragraph(v, st2))
                except Exception:
                    cells.append(Paragraph(v, td_s))
            else:
                cells.append(Paragraph(v, td_s))
        ana_rows.append(cells)

    ana_tbl = Table([ana_hdr] + ana_rows, colWidths=ana_ws,
                    hAlign="LEFT", repeatRows=1)
    ana_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  NAVY),
        ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 3),
        ("RIGHTPADDING",  (0,0), (-1,-1), 3),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [rl.white, LGRAY]),
        ("REPEATROWS",    (0,0), (0,0)),
        ("LINEBELOW",     (0,-1),(-1,-1), 1.5, BLUE),
    ]))
    els.append(ana_tbl)

    # ══════════════════════════════════
    # § 6  REPORT VERIFICATION
    # ══════════════════════════════════

    dept_s = department if department else "the Department"
    ver_para = Paragraph(
        f"This is an officially verified result analysis document of the "
        f"{dept_s}, Carmel College of Engineering and Technology.",
        ver_body,
    )

    if qr_path and os.path.exists(qr_path):
        ver_qr = Table(
            [[Image(qr_path, width=58, height=58)],
             [Paragraph("<para align='center'><font size='6.5' color='#64748B'>"
                        "Scan to verify</font></para>", S["Normal"])]],
            colWidths=[AW * 0.26],
        )
        ver_qr.setStyle(TableStyle([
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("LEFTPADDING",   (0,0), (-1,-1), 0),
            ("RIGHTPADDING",  (0,0), (-1,-1), 0),
            ("TOPPADDING",    (0,0), (-1,-1), 0),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
        ]))
    else:
        ver_qr = Paragraph("", S["Normal"])

    # Signature: 2-row inner table — spacer row + label row with LINEABOVE
    sig_tbl = Table(
        [[""],
         [Paragraph("Head of Department", sig_lbl)]],
        colWidths=[AW * 0.30],
        rowHeights=[30, None],
    )
    sig_tbl.setStyle(TableStyle([
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (0,0),   "BOTTOM"),
        ("VALIGN",        (0,1), (0,1),   "TOP"),
        ("LINEABOVE",     (0,1), (-1,1),  0.8, DARK),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ("RIGHTPADDING",  (0,0), (-1,-1), 10),
        ("TOPPADDING",    (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,1), (-1,1),  4),
    ]))

    ver_outer = Table(
        [[ver_para, ver_qr, sig_tbl]],
        colWidths=[AW * 0.44, AW * 0.26, AW * 0.30],
    )
    ver_outer.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",   (0,0), (0,0),   4),
        ("LEFTPADDING",   (1,0), (-1,0),  0),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
        ("TOPPADDING",    (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
    ]))

    els.append(KeepTogether([
        Spacer(1, 22),
        HRFlowable(width="100%", thickness=1, color=BLUE,
                   spaceBefore=0, spaceAfter=6),
        Paragraph("Report Verification", ver_head),
        ver_outer,
        HRFlowable(width="100%", thickness=1, color=BLUE,
                   spaceBefore=4, spaceAfter=0),
    ]))

    doc.build(els, onFirstPage=_footer, onLaterPages=_footer)
    return output_path


# ═══════════════════════════════════════════════════════════════
# EXCEL REPORT
# ═══════════════════════════════════════════════════════════════

def generate_excel_report(
    df, analysis_df, output_path, report_id, qr_path,
    department="", semester="", batch=""
):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df = df.fillna("")

    wb = Workbook()
    ws = wb.active
    ws.title = "Result Report"

    date_str  = datetime.now().strftime("%d %B %Y")
    sem_lbl   = f"S{semester}" if semester else str(semester)
    subj_cols = [c for c in df.columns if c not in ("Register No", "SGPA", "Arrear")]

    # ── Thin border helper ────────────────────────────────────
    THIN = Side(style="thin", color="CBD5E1")
    MED  = Side(style="medium", color="1E3A8A")

    def bdr(left=THIN, right=THIN, top=THIN, bottom=THIN):
        return Border(left=left, right=right, top=top, bottom=bottom)

    def cell(row, col, val, bold=False, size=9, color="0F172A",
             bg=None, align="left", va="center", wrap=False, border=True,
             italic=False):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=size, bold=bold,
                           color=color, italic=italic)
        c.alignment = Alignment(horizontal=align, vertical=va,
                                wrap_text=wrap)
        if bg:     c.fill   = PatternFill("solid", fgColor=bg.lstrip("#"))
        if border: c.border = bdr()
        return c

    # Number of data columns = subjects + reg + sgpa + arrear
    NC = len(subj_cols)          # subject count
    TC = 1 + NC + 1 + 1         # Register No + subjects + SGPA + Arrear

    r = 1  # current row pointer

    # ══════════════════════════════════════════════════════════
    # BLOCK 1 — DOCUMENT HEADER  (no borders, no fill — clean)
    # ══════════════════════════════════════════════════════════

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    c0 = ws.cell(row=r, column=1,
                 value="CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY")
    c0.font      = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    c0.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    c1 = ws.cell(row=r, column=1, value="RESULT ANALYSIS REPORT")
    c1.font      = Font(name="Calibri", size=11, bold=True, color="DC2626")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    # Thin separator below title
    for col in range(1, TC + 1):
        ws.cell(row=r, column=col).border = Border(
            bottom=Side(style="thin", color="CBD5E1"))
    ws.row_dimensions[r].height = 3
    r += 1

    # ── Meta: key | : | value ────────────────────────────────
    meta_pairs = [
        ("Report ID",  report_id),
        ("Date",       date_str),
        ("Department", department),
        ("Semester",   sem_lbl),
    ]
    if batch:
        meta_pairs.append(("Batch", batch))

    meta_start = r
    for key, val in meta_pairs:
        # key (cols 1-2, right-aligned bold)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ck = ws.cell(row=r, column=1, value=key)
        ck.font      = Font(name="Calibri", size=9, bold=True, color="0F172A")
        ck.alignment = Alignment(horizontal="right", vertical="center")

        # colon (col 3)
        cc = ws.cell(row=r, column=3, value=":")
        cc.font      = Font(name="Calibri", size=9, color="0F172A")
        cc.alignment = Alignment(horizontal="center", vertical="center")

        # value (cols 4 to TC)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=TC)
        cv = ws.cell(row=r, column=4, value=val)
        cv.font      = Font(name="Calibri", size=9, color="0F172A")
        cv.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)

        ws.row_dimensions[r].height = 14
        r += 1

    # Bottom border of meta block
    for col in range(1, TC + 1):
        existing = ws.cell(row=r - 1, column=col).border
        ws.cell(row=r - 1, column=col).border = Border(
            left=existing.left, right=existing.right, top=existing.top,
            bottom=Side(style="medium", color="3B82F6"))
    r += 1  # one blank gap row
    ws.row_dimensions[r - 1].height = 4

    # QR image top-right
    if qr_path and os.path.exists(qr_path):
        try:
            qr_img        = XLImage(qr_path)
            qr_img.width  = 65
            qr_img.height = 65
            ws.add_image(qr_img, f"{get_column_letter(TC - 1)}1")
        except Exception:
            pass

    # ══════════════════════════════════════════════════════════
    # BLOCK 2 — RESULT OVERVIEW
    # Simple two-column table: label | value
    # ══════════════════════════════════════════════════════════

    sgpas    = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa = round(sum(sgpas) / len(sgpas), 2) if sgpas else 0
    max_sgpa = round(max(sgpas), 2) if sgpas else 0
    min_sgpa = round(min(sgpas), 2) if sgpas else 0
    failures = sum(1 for s in sgpas if s == 0)

    # Section label row (no fill, just bold underlined text)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl = ws.cell(row=r, column=1, value="RESULT OVERVIEW")
    sl.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl.alignment = Alignment(horizontal="left", vertical="center")
    sl.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    LBL_W = 5   # label spans cols 1-5
    VAL_W = 3   # value spans cols 6-8
    overview = [
        ("Total Students",    len(df),     "0F172A"),
        ("Total Subjects",    len(subj_cols), "0F172A"),
        ("Average SGPA",      avg_sgpa,    "0F172A"),
        ("Highest SGPA",      max_sgpa,    "16A34A"),
        ("Lowest SGPA",       min_sgpa,    "0F172A"),
        ("Failures (SGPA 0)", failures,    "DC2626"),
    ]
    for i, (lbl, val, vc) in enumerate(overview):
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=LBL_W)
        cl = ws.cell(row=r, column=1, value=lbl)
        cl.font      = Font(name="Calibri", size=9.5, bold=True, color="0F172A")
        cl.fill      = PatternFill("solid", fgColor=bg.lstrip("#"))
        cl.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        cl.border    = bdr()

        ws.merge_cells(start_row=r, start_column=LBL_W + 1,
                       end_row=r, end_column=LBL_W + VAL_W)
        cv = ws.cell(row=r, column=LBL_W + 1, value=val)
        cv.font      = Font(name="Calibri", size=9.5, bold=True, color=vc)
        cv.fill      = PatternFill("solid", fgColor=bg.lstrip("#"))
        cv.alignment = Alignment(horizontal="center", vertical="center")
        cv.border    = bdr()
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1  # gap
    ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════════════════════════════
    # BLOCK 3 — STUDENTS RESULT
    # ══════════════════════════════════════════════════════════

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl2 = ws.cell(row=r, column=1, value="STUDENTS RESULT")
    sl2.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl2.alignment = Alignment(horizontal="left", vertical="center")
    sl2.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    # Header row
    for ci, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=r, column=ci, value=col_name)
        c.font      = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", wrap_text=True,
                                indent=1 if ci == 1 else 0)
        c.border    = bdr()
    ws.row_dimensions[r].height = 18
    r += 1

    for _, row_data in df.iterrows():
        sgpa   = safe_float(row_data.get("SGPA", 0))
        row_bg = ("FFEBEB" if sgpa == 0
                  else ("FFFBEB" if 0 < sgpa < 5.0 else "FFFFFF"))

        for ci, col_name in enumerate(df.columns, 1):
            val = str(row_data[col_name]).strip()
            val = "" if val in ("nan", "None") else val

            if col_name == "Register No":
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, bold=True,
                                   color="DC2626" if sgpa == 0 else "0F172A")
                c.fill      = PatternFill("solid", fgColor=row_bg.lstrip("#"))
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1)
                c.border    = bdr()

            elif col_name in subj_cols:
                bg_h, fg_h = _grade_colors(val.upper())
                c = ws.cell(row=r, column=ci, value=val if val else "")
                if bg_h and val:
                    c.fill  = PatternFill("solid", fgColor=bg_h.lstrip("#"))
                    c.font  = Font(name="Calibri", size=8, bold=True,
                                   color=fg_h.lstrip("#"))
                else:
                    c.fill  = PatternFill("solid", fgColor=row_bg.lstrip("#"))
                    c.font  = Font(name="Calibri", size=8, color="0F172A")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

            elif col_name == "SGPA":
                fg = ("DC2626" if sgpa == 0
                      else ("16A34A" if sgpa >= 7.5 else "0F172A"))
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, bold=True, color=fg)
                c.fill      = PatternFill("solid", fgColor=row_bg.lstrip("#"))
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

            else:  # Arrear
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, color="0F172A")
                c.fill      = PatternFill("solid", fgColor=row_bg.lstrip("#"))
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

        ws.row_dimensions[r].height = 13
        r += 1

    r += 1
    ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════════════════════════════
    # BLOCK 4 — TOP PERFORMERS
    # ══════════════════════════════════════════════════════════

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl3 = ws.cell(row=r, column=1, value="TOP PERFORMERS")
    sl3.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl3.alignment = Alignment(horizontal="left", vertical="center")
    sl3.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    # Header: 3 simple columns, no merging, align with data table
    for ci, (hdr, w) in enumerate(
            [("Rank", 1), ("Register No", 5), ("SGPA", 3)], 0):
        col_start = 1 + sum(x[1] for x in
                            [("Rank", 1), ("Register No", 5), ("SGPA", 3)][:ci])
        col_end   = col_start + w - 1
        ws.merge_cells(start_row=r, start_column=col_start,
                       end_row=r, end_column=col_end)
        c = ws.cell(row=r, column=col_start, value=hdr)
        c.font      = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="D97706")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    ws.row_dimensions[r].height = 16
    r += 1

    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)

    for rank, (_, rd) in enumerate(top_df.iterrows(), 1):
        bg = "FFFBEB" if rank % 2 else "FFFFFF"
        cols = [(1, 1, str(rank), "center"),
                (3, 7, rd["Register No"], "left"),
                (8, 10, str(rd["SGPA"]), "center")]
        for c1, c2, val, al in cols:
            ws.merge_cells(start_row=r, start_column=c1,
                           end_row=r, end_column=c2)
            c = ws.cell(row=r, column=c1, value=val)
            c.font      = Font(name="Calibri", size=9,
                               bold=(c1 == 1 or c1 == 8),
                               color="D97706" if c1 == 1 else
                                     ("16A34A" if c1 == 8 else "0F172A"))
            c.fill      = PatternFill("solid", fgColor=bg.lstrip("#"))
            c.alignment = Alignment(horizontal=al, vertical="center",
                                    indent=1 if al == "left" else 0)
            c.border    = bdr()
        ws.row_dimensions[r].height = 14
        r += 1

    r += 1
    ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════════════════════════════
    # BLOCK 5 — SUBJECT PERFORMANCE ANALYSIS
    # ══════════════════════════════════════════════════════════

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl4 = ws.cell(row=r, column=1, value="SUBJECT PERFORMANCE ANALYSIS")
    sl4.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl4.alignment = Alignment(horizontal="left", vertical="center")
    sl4.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    ana_cols = list(analysis_df.columns)
    for ci, h in enumerate(ana_cols, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font      = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", wrap_text=True,
                                indent=1 if ci == 1 else 0)
        c.border    = bdr()
    ws.row_dimensions[r].height = 20
    r += 1

    for i, (_, rd) in enumerate(analysis_df.iterrows()):
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(ana_cols, 1):
            val = str(rd[col])
            c   = ws.cell(row=r, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg.lstrip("#"))
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center",
                indent=1 if ci == 1 else 0)
            c.border    = bdr()

            if ci == 1:
                c.font = Font(name="Calibri", size=8, bold=True,
                              color="0F172A")
            elif "%" in val:
                try:
                    pct = float(val.replace("%", ""))
                    fg  = ("DC2626" if pct < 60
                           else ("D97706" if pct < 80 else "16A34A"))
                    c.font = Font(name="Calibri", size=8, bold=True, color=fg)
                except Exception:
                    c.font = Font(name="Calibri", size=8, color="0F172A")
            else:
                c.font = Font(name="Calibri", size=8, color="0F172A")

        ws.row_dimensions[r].height = 13
        r += 1

    r += 1
    ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════════════════════════════
    # BLOCK 6 — REPORT VERIFICATION
    # ══════════════════════════════════════════════════════════

    dept_s   = department if department else "the Department"
    ver_text = (f"This is an officially verified result analysis document "
                f"of the {dept_s}, Carmel College of Engineering "
                f"and Technology.")

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl5 = ws.cell(row=r, column=1, value="REPORT VERIFICATION")
    sl5.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl5.alignment = Alignment(horizontal="left", vertical="center")
    sl5.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    half = max(TC // 2, 5)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=half)
    cv = ws.cell(row=r, column=1, value=ver_text)
    cv.font      = Font(name="Calibri", size=9, color="0F172A")
    cv.alignment = Alignment(horizontal="left", vertical="center",
                             wrap_text=True, indent=1)
    cv.border    = bdr()

    sig_col = half + 2
    if sig_col <= TC:
        ws.merge_cells(start_row=r + 2, start_column=sig_col,
                       end_row=r + 2, end_column=TC)
        sg = ws.cell(row=r + 2, column=sig_col, value="Head of Department")
        sg.font      = Font(name="Calibri", size=9, bold=True, color="0F172A")
        sg.alignment = Alignment(horizontal="center", vertical="top")
        sg.border    = Border(top=Side(style="thin", color="0F172A"))

    for rr in range(r, r + 4):
        ws.row_dimensions[rr].height = 16

    # ══════════════════════════════════════════════════════════
    # COLUMN WIDTHS  (fit content, no wasted space)
    # ══════════════════════════════════════════════════════════

    ws.column_dimensions["A"].width = 15   # Register No / label
    ws.column_dimensions["B"].width = 4    # meta col-2 (key partner)
    ws.column_dimensions["C"].width = 2.5  # colon col
    # Subject columns: cols 4 to TC-2 (dynamic)
    for ci in range(4, TC - 1):
        ws.column_dimensions[get_column_letter(ci)].width = 6.5
    ws.column_dimensions[get_column_letter(TC - 1)].width = 5.5  # SGPA
    ws.column_dimensions[get_column_letter(TC)].width     = 5.5  # Arrear

    # Freeze header + meta rows so column labels stay visible while scrolling
    freeze_r = meta_start + len(meta_pairs) + 2
    ws.freeze_panes = f"B{freeze_r}"

    # Print settings
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{meta_start + len(meta_pairs)}"

    wb.save(output_path)
    return output_path

    # ── Colour / border shortcuts ─────────────────────────────
    def _c(h):  return h.lstrip("#")
    def _f(h):  return _fill(h)
    def _b():   return _border()

    def _cell(row, col, val, bold=False, size=9, color="0F172A",
              bg=None, align="center", va="center", wrap=False, border=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, size=size, color=color, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical=va, wrap_text=wrap)
        if bg:     c.fill   = _f(bg)
        if border: c.border = _b()
        return c

    def _span(r1, c1, r2, c2, val="", bold=False, size=9, color="0F172A",
              bg=None, align="center", wrap=False, border=False):
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1,
                           end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=val)
        c.font      = Font(bold=bold, size=size, color=color, name="Calibri")
        c.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
        if bg:     c.fill   = _f(bg)
        if border: c.border = _b()
        return c

    def _section(row, text, color=_NAVY):
        _span(row, 1, row, TC, val=f"  {text}",
              bold=True, size=10, color="FFFFFF",
              bg=color, align="left", border=False)
        ws.cell(row=row, column=1).border = _b()
        ws.row_dimensions[row].height = 18

    r = 1

    # ══════════════════════════════════
    # § 1  HEADER
    # ══════════════════════════════════

    # College name bar
    _span(r, 1, r, TC, val="CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY",
          bold=True, size=13, color="FFFFFF", bg=_NAVY, align="center")
    ws.row_dimensions[r].height = 26
    r += 1

    # Report title bar
    _span(r, 1, r, TC, val="RESULT ANALYSIS REPORT",
          bold=True, size=11, color=_c(_RED), bg="EFF6FF", align="center")
    ws.row_dimensions[r].height = 20
    r += 1

    # ── Meta block: col1-2 = label (right), col3 = colon, col4+ = value ──
    # Fixed widths: label=2 cols, sep=1 col, value = rest
    meta_pairs = [
        ("Report ID",  report_id),
        ("Date",       date_str),
        ("Department", department),
        ("Semester",   sem_lbl),
    ]
    if batch:
        meta_pairs.append(("Batch", batch))

    meta_start = r
    for key, val in meta_pairs:
        _span(r, 1, r, 2, val=key, bold=True, size=9, color="0F172A",
              align="right", border=False)
        ws.cell(r, 3).value = ":"
        ws.cell(r, 3).font  = Font(size=9, name="Calibri", color="0F172A")
        ws.cell(r, 3).alignment = Alignment(horizontal="center", vertical="center")
        _span(r, 4, r, TC, val=val, size=9, color="0F172A",
              align="left", wrap=True, border=False)
        ws.row_dimensions[r].height = 15
        r += 1

    # Blue separator
    _span(r, 1, r, TC, bg=_c(_BLUE))
    ws.row_dimensions[r].height = 2
    r += 1

    # QR image (top-right, overlapping header rows)
    if qr_path and os.path.exists(qr_path):
        try:
            qr_img        = XLImage(qr_path)
            qr_img.width  = 72
            qr_img.height = 72
            qr_col        = get_column_letter(max(1, TC - 1))
            ws.add_image(qr_img, f"{qr_col}1")
        except Exception:
            pass

    # ══════════════════════════════════
    # § 2  RESULT OVERVIEW
    # ══════════════════════════════════

    sgpas    = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa = round(sum(sgpas) / len(sgpas), 2) if sgpas else 0
    max_sgpa = round(max(sgpas), 2) if sgpas else 0
    min_sgpa = round(min(sgpas), 2) if sgpas else 0
    failures = sum(1 for s in sgpas if s == 0)

    r += 1  # one blank row gap
    _section(r, "RESULT OVERVIEW")
    r += 1

    # Overview table: label spans cols 1-6, value spans 7-10
    LW, VW = 6, 4   # label width in cols, value width in cols
    overview = [
        ("Total Students",    len(df),     None),
        ("Total Subjects",    len(subj_cols), None),
        ("Average SGPA",      avg_sgpa,    None),
        ("Highest SGPA",      max_sgpa,    "16A34A"),
        ("Lowest SGPA",       min_sgpa,    None),
        ("Failures (SGPA 0)", failures,    "DC2626"),
    ]
    for i, (label, val, val_col) in enumerate(overview):
        bg = "F1F5F9" if i % 2 == 0 else "FFFFFF"
        _span(r, 1, r, LW, val=label, bold=True, size=9.5, color="0F172A",
              bg=bg, align="left", border=True)
        ws.cell(r, 1).alignment = Alignment(
            horizontal="left", vertical="center", indent=1)
        _span(r, LW + 1, r, LW + VW, val=val, bold=True, size=9.5,
              color=val_col if val_col else "0F172A",
              bg=bg, align="center", border=True)
        ws.row_dimensions[r].height = 17
        r += 1

    # ══════════════════════════════════
    # § 3  STUDENTS RESULT
    # ══════════════════════════════════

    r += 1
    _section(r, "STUDENTS RESULT")
    r += 1

    for ci, col_name in enumerate(df.columns, 1):
        _cell(r, ci, col_name, bold=True, size=8, color="FFFFFF",
              bg=_c(_NAVY), align="center", wrap=True)
    ws.row_dimensions[r].height = 20
    r += 1

    for _, row_data in df.iterrows():
        sgpa   = safe_float(row_data.get("SGPA", 0))
        row_bg = ("FFEBEB" if sgpa == 0
                  else ("FFFBEB" if 0 < sgpa < 5.0 else "FFFFFF"))

        for ci, col_name in enumerate(df.columns, 1):
            val    = str(row_data[col_name]).strip()
            val    = "" if val in ("nan", "None") else val
            is_reg = col_name == "Register No"
            is_sub = col_name in subj_cols
            is_sgp = col_name == "SGPA"

            if is_reg:
                _cell(r, ci, val, bold=True, size=8,
                      color="DC2626" if sgpa == 0 else "0F172A",
                      bg=row_bg, align="left")
            elif is_sub:
                bg_h, fg_h = _grade_colors(val.upper())
                if bg_h:
                    _cell(r, ci, val, bold=True, size=8,
                          color=_c(fg_h), bg=_c(bg_h))
                else:
                    _cell(r, ci, val, size=8, bg=row_bg)
            elif is_sgp:
                fg = ("DC2626" if sgpa == 0
                      else ("16A34A" if sgpa >= 7.5 else "0F172A"))
                _cell(r, ci, val, bold=True, size=8, color=fg, bg=row_bg)
            else:
                _cell(r, ci, val, size=8, bg=row_bg)

        ws.row_dimensions[r].height = 14
        r += 1

    # ══════════════════════════════════
    # § 4  TOP PERFORMERS
    # ══════════════════════════════════

    r += 1
    _section(r, "TOP PERFORMERS", color="D97706")
    r += 1

    # Header: Rank (2 cols) | Register No (5 cols) | SGPA (3 cols)
    _span(r, 1, r, 2,  val="Rank",        bold=True, size=9, color="FFFFFF",
          bg="D97706", align="center", border=True)
    _span(r, 3, r, 7,  val="Register No", bold=True, size=9, color="FFFFFF",
          bg="D97706", align="center", border=True)
    _span(r, 8, r, 10, val="SGPA",        bold=True, size=9, color="FFFFFF",
          bg="D97706", align="center", border=True)
    ws.row_dimensions[r].height = 18
    r += 1

    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)

    medals = {1: "D97706", 2: "6B7280", 3: "92400E"}  # gold / silver / bronze
    for rank, (_, rd) in enumerate(top_df.iterrows(), 1):
        bg = "FFFBEB" if rank % 2 else "FFFFFF"
        m  = medals.get(rank, "0F172A")
        _span(r, 1, r, 2, val=rank, bold=True, size=9,
              color=m, bg=bg, align="center", border=True)
        _span(r, 3, r, 7, val=rd["Register No"], size=9,
              color="0F172A", bg=bg, align="left", border=True)
        ws.cell(r, 3).alignment = Alignment(horizontal="left", vertical="center",
                                             indent=1)
        _span(r, 8, r, 10, val=rd["SGPA"], bold=True, size=9,
              color="16A34A", bg=bg, align="center", border=True)
        ws.row_dimensions[r].height = 16
        r += 1

    # ══════════════════════════════════
    # § 5  SUBJECT PERFORMANCE ANALYSIS
    # ══════════════════════════════════

    r += 1
    _section(r, "SUBJECT PERFORMANCE ANALYSIS")
    r += 1

    ana_cols = list(analysis_df.columns)
    for ci, h in enumerate(ana_cols, 1):
        _cell(r, ci, h, bold=True, size=8, color="FFFFFF",
              bg=_c(_NAVY), wrap=True)
    ws.row_dimensions[r].height = 22
    r += 1

    for i, (_, rd) in enumerate(analysis_df.iterrows()):
        bg = "F1F5F9" if i % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(ana_cols, 1):
            val = str(rd[col])
            al  = "left" if col == "Subject" else "center"
            c   = _cell(r, ci, val, size=8, color="0F172A", bg=bg, align=al)
            if col == "Subject":
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1)
            if "%" in val:
                try:
                    pct = float(val.replace("%", ""))
                    fg  = ("DC2626" if pct < 60
                           else ("D97706" if pct < 80 else "16A34A"))
                    c.font = Font(bold=True, size=8, name="Calibri", color=fg)
                except Exception:
                    pass
        ws.row_dimensions[r].height = 14
        r += 1

    # ══════════════════════════════════
    # § 6  VERIFICATION
    # ══════════════════════════════════

    r += 1
    _section(r, "REPORT VERIFICATION")
    r += 1

    dept_s   = department if department else "the Department"
    ver_text = (f"This is an officially verified result analysis document "
                f"of the {dept_s}, Carmel College of Engineering and Technology.")

    # Verification text: cols 1 to TC-3
    val_end = max(TC - 3, 4)
    _span(r, 1, r + 3, val_end, val=ver_text, size=9, color="0F172A",
          align="left", wrap=True, border=True)
    ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="top",
                                         wrap_text=True, indent=1)

    # Signature line: bottom-right
    sig_r = r + 3
    sig_c = val_end + 2
    if sig_c <= TC:
        _span(sig_r, sig_c, sig_r, TC, val="Head of Department",
              bold=True, size=9, color="0F172A",
              align="center", border=True)
    for rr in range(r, r + 5):
        ws.row_dimensions[rr].height = 16

    # ══════════════════════════════════
    # Column widths
    # ══════════════════════════════════

    ws.column_dimensions[get_column_letter(1)].width = 16   # Register No / label
    ws.column_dimensions[get_column_letter(2)].width = 4    # narrow (meta colon partner)
    ws.column_dimensions[get_column_letter(3)].width = 3    # colon col
    for ci in range(4, TC + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 7
    # SGPA and Arrear last two cols can be a touch narrower
    ws.column_dimensions[get_column_letter(TC)].width     = 6
    ws.column_dimensions[get_column_letter(TC - 1)].width = 6

    # Freeze at B(first data row) so header + meta stay fixed
    ws.freeze_panes = f"B{meta_start}"

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════
# HTML REPORT  (self-contained, print-ready)
# ═══════════════════════════════════════════════════════════════

def generate_html_report(
    df, analysis_df, report_id, department, semester, batch, qr_path
):
    date_str  = datetime.now().strftime("%d %B %Y")
    sem_lbl   = f"S{semester}" if semester else str(semester)
    qr_uri    = _qr_b64(qr_path)
    subj_cols = [c for c in df.columns if c not in ("Register No","SGPA","Arrear")]

    sgpas    = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa = round(sum(sgpas)/len(sgpas), 2) if sgpas else 0
    max_sgpa = round(max(sgpas), 2) if sgpas else 0
    min_sgpa = round(min(sgpas), 2) if sgpas else 0
    failures = sum(1 for s in sgpas if s == 0)

    # Top performers
    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)

    # ── Grade cell HTML helper ────────────────────────────────
    def _gcell(grade):
        g = str(grade).strip().upper()
        bg, fg = _grade_colors(g)
        disp = grade if grade not in ("", "nan", "None") else ""
        if bg and disp:
            return (f'<td style="background:{bg};color:{fg};'
                    f'font-weight:700;text-align:center;padding:3px 2px;">'
                    f'{disp}</td>')
        return f'<td style="text-align:center;padding:3px 2px;">{disp}</td>'

    # ── Student rows ──────────────────────────────────────────
    student_rows = ""
    for _, row in df.iterrows():
        sgpa   = safe_float(row.get("SGPA", 0))
        row_bg = ("#FFEBEB" if sgpa == 0
                  else ("#FFFBEB" if 0 < sgpa < 5.0 else "#FFFFFF"))
        reg    = row.get("Register No", "")
        reg_c  = "#DC2626" if sgpa == 0 else "#0F172A"
        sgpa_c = ("#DC2626" if sgpa == 0
                  else ("#16A34A" if sgpa >= 7.5 else "#0F172A"))

        cells = (f'<td style="background:{row_bg};color:{reg_c};'
                 f'font-weight:700;padding:4px 6px;white-space:nowrap;">{reg}</td>')
        for col in subj_cols:
            cells += _gcell(str(row.get(col, "")))
        cells += (f'<td style="background:{row_bg};color:{sgpa_c};'
                  f'font-weight:700;text-align:center;padding:4px;">'
                  f'{row.get("SGPA","")}</td>')
        cells += (f'<td style="background:{row_bg};text-align:center;'
                  f'padding:4px;">{row.get("Arrear","")}</td>')
        student_rows += f"<tr>{cells}</tr>\n"

    # ── Subject analysis rows ─────────────────────────────────
    ana_cols_list = list(analysis_df.columns)
    ana_rows = ""
    for i, (_, row) in enumerate(analysis_df.iterrows()):
        bg = "#F1F5F9" if i % 2 == 0 else "#FFFFFF"
        cells = ""
        for col in ana_cols_list:
            v = str(row[col])
            align = "left" if col == "Subject" else "center"
            style = f"background:{bg};padding:5px;text-align:{align};"
            if "%" in v:
                try:
                    pct = float(v.replace("%",""))
                    fc = ("#DC2626" if pct < 60
                          else ("#D97706" if pct < 80 else "#16A34A"))
                    style += f"color:{fc};font-weight:700;"
                except Exception:
                    pass
            cells += f'<td style="{style}">{v}</td>'
        ana_rows += f"<tr>{cells}</tr>\n"

    # ── Top performers rows ───────────────────────────────────
    top_rows = ""
    for rank, (_, row) in enumerate(top_df.iterrows(), 1):
        bg = "#FFFBEB" if rank % 2 else "#FFFFFF"
        top_rows += (f'<tr style="background:{bg};">'
                     f'<td style="text-align:center;font-weight:700;'
                     f'color:#D97706;padding:6px;">{rank}</td>'
                     f'<td style="padding:6px;">{row["Register No"]}</td>'
                     f'<td style="text-align:center;font-weight:700;'
                     f'padding:6px;">{row["SGPA"]}</td></tr>\n')

    # ── Subject header cells ──────────────────────────────────
    subj_th = "".join(
        f'<th style="padding:5px 3px;white-space:nowrap;">{s}</th>'
        for s in subj_cols)
    ana_th = "".join(
        f'<th style="padding:5px 3px;">{h}</th>'
        for h in ana_cols_list)

    qr_block = (f'<img src="{qr_uri}" style="width:80px;height:80px;" alt="QR"/>'
                if qr_uri else
                '<div style="width:80px;height:80px;border:1px dashed #cbd5e1;'
                'border-radius:4px;"></div>')

    dept_s = department if department else "the Department"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Result Analysis Report — {department}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',Helvetica,Arial,sans-serif;
     color:#0F172A;background:#f8fafc;font-size:12px;}}
.wrap{{max-width:210mm;margin:0 auto;background:#fff;
       box-shadow:0 1px 8px rgba(0,0,0,.12);}}

/* Header */
.hdr-college{{background:#1E3A8A;color:#fff;text-align:center;
              padding:12px 16px;font-size:15px;font-weight:700;
              letter-spacing:.3px;}}
.hdr-title{{background:#EFF6FF;text-align:center;
            padding:7px 16px;font-size:13px;font-weight:700;
            color:#DC2626;border-bottom:2.5px solid #3B82F6;}}
.hdr-meta{{display:flex;justify-content:space-between;
           align-items:flex-start;padding:12px 16px 10px;
           border-bottom:2px solid #3B82F6;}}
.meta-table{{border-collapse:collapse;font-size:9pt;}}
.meta-table td{{padding:2px 4px;vertical-align:top;}}
.meta-key{{font-weight:700;text-align:right;white-space:nowrap;}}
.meta-sep{{text-align:center;padding:2px 6px;}}
.qr-area{{text-align:center;}}
.qr-area p{{font-size:8px;color:#64748B;margin-top:4px;max-width:90px;}}

/* Sections */
.body{{padding:0 16px 16px;}}
.sect-title{{font-size:11px;font-weight:700;color:#0F172A;
             border-bottom:1.5px solid #3B82F6;
             padding:10px 0 4px;margin-top:14px;margin-bottom:8px;}}

/* Overview */
.ov-table{{width:52%;border-collapse:collapse;}}
.ov-table td{{border:.5px solid #CBD5E1;padding:6px 10px;
              font-size:9.5pt;}}
.ov-table tr:nth-child(even){{background:#F1F5F9;}}
.ov-key{{font-weight:700;}}
.ov-val{{text-align:center;font-weight:700;}}

/* Result table */
.rtbl{{width:100%;border-collapse:collapse;font-size:8pt;}}
.rtbl th{{background:#1E3A8A;color:#fff;padding:5px 3px;
          text-align:center;font-weight:600;white-space:nowrap;}}
.rtbl td{{border:.35px solid #CBD5E1;}}
.rtbl tr:nth-child(even) td:not([style*="background"]){{background:#F8FAFC;}}

/* Top performers */
.top-tbl{{width:55%;border-collapse:collapse;font-size:9pt;}}
.top-tbl th{{background:#D97706;color:#fff;padding:6px;text-align:center;}}
.top-tbl td{{border:.5px solid #CBD5E1;}}

/* Subject analysis */
.atbl{{width:100%;border-collapse:collapse;font-size:8pt;}}
.atbl th{{background:#1E3A8A;color:#fff;padding:5px 3px;text-align:center;}}
.atbl tr:nth-child(even){{background:#F1F5F9;}}
.atbl td{{border:.35px solid #CBD5E1;padding:4px 3px;text-align:center;}}
.atbl td:first-child{{text-align:left;padding-left:6px;}}

/* Verification */
.ver{{display:flex;align-items:center;gap:0;
      border:1px solid #E2E8F0;border-radius:6px;
      padding:12px;margin-top:16px;}}
.ver-text{{flex:1;font-size:9pt;line-height:1.5;}}
.ver-title{{font-weight:700;font-size:11pt;color:#1D4ED8;margin-bottom:6px;}}
.ver-qr{{text-align:center;padding:0 20px;}}
.ver-qr p{{font-size:7.5pt;color:#64748B;margin-top:4px;}}
.ver-sig{{flex:1;text-align:center;}}
.ver-sig .line{{border-top:1px solid #0F172A;width:75%;
                margin:28px auto 6px;}}
.ver-sig .lbl{{font-weight:700;font-size:9pt;}}

/* Print */
@media print{{
  body{{background:#fff;-webkit-print-color-adjust:exact;
        print-color-adjust:exact;}}
  .no-print{{display:none;}}
  .wrap{{box-shadow:none;max-width:100%;}}
  .page-break{{page-break-before:always;}}
  .rtbl{{font-size:7pt;}}
}}

</style>
</head>
<body>
<div class="wrap">

<!-- HEADER -->
<div class="hdr-college">CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY</div>
<div class="hdr-title">RESULT ANALYSIS REPORT</div>

<div class="hdr-meta">
  <table class="meta-table">
    <tr>
      <td class="meta-key">Report ID</td>
      <td class="meta-sep">:</td>
      <td>{report_id}</td>
    </tr>
    <tr>
      <td class="meta-key">Date</td>
      <td class="meta-sep">:</td>
      <td>{date_str}</td>
    </tr>
    <tr>
      <td class="meta-key">Department</td>
      <td class="meta-sep">:</td>
      <td>{department}</td>
    </tr>
    <tr>
      <td class="meta-key">Semester</td>
      <td class="meta-sep">:</td>
      <td>{sem_lbl}</td>
    </tr>
    {"<tr><td class='meta-key'>Batch</td><td class='meta-sep'>:</td><td>" + batch + "</td></tr>" if batch else ""}
  </table>
  <div class="qr-area">
    {qr_block}
    <p>Scan to verify<br>authenticity</p>
  </div>
</div>

<!-- BODY -->
<div class="body">

<!-- RESULT OVERVIEW -->
<div class="sect-title">RESULT OVERVIEW</div>
<table class="ov-table">
<tbody>
  <tr><td class="ov-key">Total Students</td>
      <td class="ov-val">{len(df)}</td></tr>
  <tr><td class="ov-key">Total Subjects</td>
      <td class="ov-val">{len(subj_cols)}</td></tr>
  <tr><td class="ov-key">Average SGPA</td>
      <td class="ov-val">{avg_sgpa}</td></tr>
  <tr><td class="ov-key">Highest SGPA</td>
      <td class="ov-val" style="color:#16A34A;">{max_sgpa}</td></tr>
  <tr><td class="ov-key">Lowest SGPA</td>
      <td class="ov-val">{min_sgpa}</td></tr>
  <tr><td class="ov-key">Failures (SGPA 0)</td>
      <td class="ov-val" style="color:#DC2626;">{failures}</td></tr>
</tbody>
</table>

<!-- STUDENTS RESULT -->
<div class="sect-title">STUDENTS RESULT</div>
<table class="rtbl">
<thead><tr>
  <th style="padding:5px 6px;text-align:left;">Register No</th>
  {subj_th}
  <th>SGPA</th><th>Arrear</th>
</tr></thead>
<tbody>{student_rows}</tbody>
</table>

<!-- TOP PERFORMERS -->
<div class="sect-title">TOP PERFORMERS</div>
<table class="top-tbl">
<thead><tr>
  <th>Rank</th><th>Register No</th><th>SGPA</th>
</tr></thead>
<tbody>{top_rows}</tbody>
</table>

<!-- SUBJECT PERFORMANCE ANALYSIS -->
<div class="sect-title page-break" style="margin-top:20px;">
  SUBJECT PERFORMANCE ANALYSIS</div>
<table class="atbl">
<thead><tr>{ana_th}</tr></thead>
<tbody>{ana_rows}</tbody>
</table>

<!-- REPORT VERIFICATION -->
<div class="ver">
  <div class="ver-text">
    <div class="ver-title">Report Verification</div>
    This is an officially verified result analysis document of the
    {dept_s}, Carmel College of Engineering and Technology.
  </div>
  <div class="ver-qr">
    {qr_block}
    <p>Scan to verify</p>
  </div>
  <div class="ver-sig">
    <div class="line"></div>
    <div class="lbl">Head of Department</div>
  </div>
</div>

</div><!-- /body -->
</div><!-- /wrap -->
</body>
</html>"""

    return html
