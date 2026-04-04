"""
utils/report_generator.py

Three report formats — PDF, Excel (.xlsx), HTML — all sharing one design system
and a single shared preprocessing pipeline.

FIXES APPLIED (v2)
──────────────────
1.  S.No column added as first column in every format
2.  Subject column detection fixed — uses METADATA_COLS exclusion set
3.  Course Details table added (Subject Code → Name mapping)
4.  Unified column structure via preprocess_df() called before every format
5.  Null / missing value handling hardened throughout
6.  HTML column widths and wrapping fixed  (table-layout:fixed)
7.  PDF column widths recalculated to include S.No + Name columns
8.  Subject names from actual DB data (course_map dict parameter)
9.  Top Performers section kept on safe SGPA-only logic
10. Multi-format shared preprocessing eliminates format-only bugs
11. Strict column ordering: S.No | id_col | subjects… | SGPA | Arrear

Design tokens
─────────────
Primary navy  #1E3A8A  — header bars, table headers
Accent blue   #3B82F6  — section underlines, borders
Red title     #DC2626  — subtitle
Dark text     #0F172A  — body text
Light surface #F1F5F9  — alternate rows
Mid border    #CBD5E1  — table grid

Grade colour language
─────────────────────
S / A+      deep green
A  / B+     light green
B           yellow
C+ / C      amber/orange
D           salmon
P           sky blue
F / FE      solid red
"""

import os
import base64
from datetime import datetime

# ── ReportLab ────────────────────────────────────────────────
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer, HRFlowable,
    Image, KeepTogether,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors as rl
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── openpyxl ─────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

#---------------assist file--------------------------------------
from utils.subject_utils import build_course_map

# ═══════════════════════════════════════════════════════════════
# SHARED DESIGN TOKENS
# ═══════════════════════════════════════════════════════════════

_NAVY  = "#1E3A8A"
_BLUE  = "#3B82F6"
_RED   = "#DC2626"
_DARK  = "#0F172A"
_LGRAY = "#F1F5F9"
_MGRAY = "#CBD5E1"
_WHITE = "#FFFFFF"

GRADE_COLORS = {
    "S":      (_WHITE,   "#166534", "#166534"),
    "A+":     ("#15803D", _WHITE,   "#15803D"),
    "A":      ("#86EFAC", "#14532D", "#86EFAC"),
    "B+":     ("#FEF08A", "#713F12", "#FEF08A"),
    "B":      ("#FDE047", "#713F12", "#FDE047"),
    "C+":     ("#FED7AA", "#7C2D12", "#FED7AA"),
    "C":      ("#FDBA74", "#7C2D12", "#FDBA74"),
    "D":      ("#FCA5A5", "#7F1D1D", "#FCA5A5"),
    "P":      ("#BAE6FD", "#0C4A6E", "#BAE6FD"),
    "F":      (_RED,      _WHITE,    _RED),
    "FE":     (_RED,      _WHITE,    _RED),
    "ABSENT": ("#94A3B8", "#0F172A", "#94A3B8"),
}

FAIL_GRADES = {"F", "FE", "ABSENT"}

# Columns that are NEVER subjects — strict exclusion set  (FIX #2)
METADATA_COLS = {"S.No", "Register No", "Student Name", "SGPA", "Arrear"}


# ═══════════════════════════════════════════════════════════════
# SHARED PREPROCESSING  (FIX #1 #2 #4 #5 #10 #11 #12)
# ═══════════════════════════════════════════════════════════════

def preprocess_df(df):
    """
    Standardises the student DataFrame before any report format renders it.

    Guarantees
    ----------
    - All nulls replaced with ""
    - S.No column present as column 0 (sequential 1-N)
    - subject columns = every column NOT in METADATA_COLS
    - Strict column order: S.No | id_col | subjects… | SGPA | Arrear
    - SGPA column forced to float

    Returns
    -------
    df        : cleaned, reordered DataFrame
    id_col    : "Register No" or "Student Name"
    subj_cols : ordered list of subject-code column names
    """
    import pandas as pd

    df = df.copy().fillna("")                           # FIX #5 #10

    # ── Identifier column ────────────────────────────────────
    if "Register No" in df.columns:
        id_col = "Register No"
    elif "Student Name" in df.columns:
        id_col = "Student Name"
    else:
        id_col = df.columns[0]

    # ── Strict subject detection (FIX #2) ────────────────────
    subj_cols = [c for c in df.columns if c not in METADATA_COLS]

    # ── Add / resequence S.No (FIX #1 #5) ───────────────────
    df["S.No"] = range(1, len(df) + 1)

    # ── Force SGPA numeric (FIX #10) ────────────────────────
    df["SGPA"] = df["SGPA"].apply(safe_float)

    # ── Strict column ordering (FIX #11) ────────────────────
    ordered = ["S.No", id_col] + subj_cols
    if "SGPA" in df.columns:
        ordered.append("SGPA")
    if "Arrear" in df.columns:
        ordered.append("Arrear")

    df = df[[c for c in ordered if c in df.columns]]
    return df, id_col, subj_cols


# ═══════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════

def safe_float(val) -> float:
    try:
        v = str(val).strip()
        return 0.0 if v in ("", "nan", "None", "-") else float(v)
    except (ValueError, TypeError):
        return 0.0


def _c(hex_str: str):
    return rl.HexColor(hex_str)


def _qr_b64(qr_path: str) -> str:
    if not qr_path or not os.path.exists(qr_path):
        return ""
    with open(qr_path, "rb") as f:
        return "data:image/png;base64," + base64.b64encode(f.read()).decode()


def _border_xl(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_str: str):
    return PatternFill(start_color=hex_str.lstrip("#"),
                       end_color=hex_str.lstrip("#"), fill_type="solid")


def _grade_colors(grade: str):
    """Return (bg_hex, fg_hex) for a grade string."""
    g     = str(grade).strip().upper()
    entry = GRADE_COLORS.get(g)
    if entry:
        return entry[0], entry[1]
    return None, None


def fetch_subject_names(db, subj_cols: list) -> dict:
    """
    Look up subject names for every code in subj_cols from Subject_Grade.

    Works with any db handle (PyMongo or Motor) — no Streamlit dependency.
    Returns {subject_code: subject_name}.
    Missing codes get an empty string so callers never see KeyError.
    """
    if db is None or not subj_cols:
        return {}

    docs = db["Subject_Grade"].find(
        {"subject_code": {"$in": subj_cols}},
        {"_id": 0, "subject_code": 1, "subject_name": 1},
    )
    result = {doc["subject_code"]: doc.get("subject_name", "") for doc in docs}

    # Ensure every code has an entry (even if not in DB)
    for code in subj_cols:
        result.setdefault(code, "")

    return result


# ═══════════════════════════════════════════════════════════════
# PDF REPORT   (FIX #1 #2 #3 #5 #7 #8 #9 #10)
# ═══════════════════════════════════════════════════════════════

def generate_pdf_report(
    df, analysis_df, output_path, report_id, qr_path,
    department="", semester="", batch="",
    db=None,                    # passed through; names fetched internally
    mode=None
):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # ── Shared preprocessing (FIX #4) ────────────────────────
    df, id_col, subj_cols = preprocess_df(df)
    course_map = build_course_map(subj_cols, db)   # FIX #3 #8

    # ── Decide page orientation ───────────────────────────────
    PAGE = landscape(A4) if len(subj_cols) > 8 else A4
    LM, RM = 0.45 * inch, 0.45 * inch
    TM, BM = 0.45 * inch, 0.55 * inch
    AW = PAGE[0] - LM - RM

    # ── Colour aliases ────────────────────────────────────────
    NAVY  = _c(_NAVY); BLUE  = _c(_BLUE); RED   = _c(_RED)
    DARK  = _c(_DARK); LGRAY = _c(_LGRAY); MGRAY = _c(_MGRAY)
    GOLD  = _c("#D97706")

    def _footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(_c("#64748B"))
        canvas.drawString(LM, BM * 0.42,
            "APJ Abdul Kalam Technological University  ·  "
            "Carmel College of Engineering and Technology")
        canvas.drawRightString(PAGE[0] - RM, BM * 0.42, f"Page {doc.page}")
        if doc.page > 1:
            canvas.setStrokeColor(MGRAY)
            canvas.setLineWidth(0.4)
            canvas.line(LM, PAGE[1] - TM * 0.6, PAGE[0] - RM, PAGE[1] - TM * 0.6)
        canvas.restoreState()

    doc = SimpleDocTemplate(
        output_path, pagesize=PAGE,
        leftMargin=LM, rightMargin=RM,
        topMargin=TM, bottomMargin=BM,
    )

    # ── Paragraph styles ─────────────────────────────────────
    S = getSampleStyleSheet()

    def _ps(name, parent=None, **kw):
        return ParagraphStyle(name, parent=parent or S["Normal"], **kw)

    college_s = _ps("College", fontName="Helvetica-Bold", fontSize=13,
                    textColor=rl.white, alignment=TA_CENTER, leading=18)
    red_s     = _ps("RedTitle", fontName="Helvetica-Bold", fontSize=12,
                    textColor=RED, alignment=TA_CENTER, leading=16)
    sect_s    = _ps("Sect", fontName="Helvetica-Bold", fontSize=10.5,
                    textColor=DARK, leading=14, spaceBefore=0, spaceAfter=4)
    meta_key  = _ps("MKey", fontName="Helvetica-Bold", fontSize=8.5,
                    textColor=DARK, alignment=TA_RIGHT, leading=14)
    meta_sep  = _ps("MSep", fontName="Helvetica", fontSize=8.5,
                    textColor=DARK, alignment=TA_CENTER, leading=14)
    meta_val  = _ps("MVal", fontName="Helvetica", fontSize=8.5,
                    textColor=DARK, alignment=TA_LEFT, leading=14)
    th_s      = _ps("TH", fontName="Helvetica-Bold", fontSize=7,
                    textColor=rl.white, alignment=TA_CENTER, leading=9)
    td_s      = _ps("TD", fontName="Helvetica", fontSize=7,
                    alignment=TA_CENTER, leading=9)
    td_l      = _ps("TDL", fontName="Helvetica", fontSize=7,
                    alignment=TA_LEFT, leading=9)
    td_b      = _ps("TDB", fontName="Helvetica-Bold", fontSize=7,
                    alignment=TA_CENTER, leading=9)
    td_sno    = _ps("TSno", fontName="Helvetica", fontSize=6.5,
                    alignment=TA_CENTER, leading=9)
    ov_key    = _ps("OVK", fontName="Helvetica-Bold", fontSize=9.5,
                    textColor=DARK, alignment=TA_LEFT, leading=13)
    ov_val    = _ps("OVV", fontName="Helvetica-Bold", fontSize=9.5,
                    textColor=DARK, alignment=TA_CENTER, leading=13)
    top_rank  = _ps("Rank", fontName="Helvetica-Bold", fontSize=9,
                    textColor=GOLD, alignment=TA_CENTER, leading=13)
    ver_head  = _ps("VH", fontName="Helvetica-Bold", fontSize=11,
                    textColor=_c("#1D4ED8"), leading=16, spaceAfter=4)
    ver_body  = _ps("VB", fontName="Helvetica", fontSize=8.5,
                    textColor=DARK, leading=13)
    sig_lbl   = _ps("SL", fontName="Helvetica-Bold", fontSize=9,
                    textColor=DARK, alignment=TA_CENTER, leading=13)
    cd_code   = _ps("CDCode", fontName="Helvetica-Bold", fontSize=7.5,
                    textColor=DARK, alignment=TA_LEFT, leading=10)
    cd_name   = _ps("CDName", fontName="Helvetica", fontSize=7.5,
                    textColor=DARK, alignment=TA_LEFT, leading=10)

    def blue_rule(thick=1.2):
        return HRFlowable(width="100%", thickness=thick,
                          color=BLUE, spaceBefore=2, spaceAfter=4)

    def section(title):
        return [Spacer(1, 10), Paragraph(title, sect_s), blue_rule(0.8)]

    # ── Pre-calculate ─────────────────────────────────────────
    sgpas    = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa = round(sum(sgpas) / len(sgpas), 2) if sgpas else 0
    max_sgpa = round(max(sgpas), 2) if sgpas else 0
    min_sgpa = round(min(sgpas), 2) if sgpas else 0
    failures = sum(1 for s in sgpas if s == 0)
    date_str = datetime.now().strftime("%d %B %Y")
    sem_lbl  = f"S{semester}" if semester else ""

    els = []

    # ══════════════════════════════════
    # § 1  HEADER
    # ══════════════════════════════════
    hdr_bar = Table(
        [[Paragraph("CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY", college_s)]],
        colWidths=[AW],
    )
    hdr_bar.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), NAVY),
        ("TOPPADDING",    (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
    ]))
    els.append(hdr_bar)

    sub_bar = Table([[Paragraph("RESULT ANALYSIS REPORT", red_s)]], colWidths=[AW])
    sub_bar.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), _c("#EFF6FF")),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LINEBELOW",     (0,0), (-1,-1), 2, BLUE),
    ]))
    els.append(sub_bar)
    els.append(Spacer(1, 6))

    # Meta + QR
    LKEY, LSEP = 62, 12
    LVAL = AW * 0.65 - LKEY - LSEP
    meta_id_s = ParagraphStyle("MID", parent=meta_val, fontSize=7.5, leading=11)

    def _mrow(key, val, vs=None):
        return [Paragraph(key, meta_key), Paragraph(":", meta_sep),
                Paragraph(val, vs or meta_val)]

    meta_rows = [
        _mrow("Report ID",  report_id, meta_id_s),
        _mrow("Date",       date_str),
        _mrow("Department", department),
        _mrow("Semester",   sem_lbl),
    ]
    if batch:
        meta_rows.append(_mrow("Batch", batch))

    meta_tbl = Table(meta_rows, colWidths=[LKEY, LSEP, LVAL])
    meta_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",   (0,0), (-1,-1), 0),
        ("TOPPADDING",    (0,0), (-1,-1), 2),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))

    if qr_path and os.path.exists(qr_path):
        qr_tbl = Table(
            [[Image(qr_path, width=68, height=68)],
             [Paragraph("<para align='center'><font size='6.5' color='#64748B'>"
                        "Scan to verify<br/>authenticity</font></para>",
                        S["Normal"])]],
            colWidths=[AW * 0.32],
        )
        qr_tbl.setStyle(TableStyle([
            ("ALIGN",  (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
    else:
        qr_tbl = Paragraph("", S["Normal"])

    info_row = Table([[meta_tbl, qr_tbl]], colWidths=[AW * 0.68, AW * 0.32])
    info_row.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
    ]))
    els.append(info_row)
    els.append(HRFlowable(width="100%", thickness=2, color=BLUE,
                          spaceBefore=0, spaceAfter=14))

    # ══════════════════════════════════
    # § 2  RESULT OVERVIEW
    # ══════════════════════════════════
    els += section("RESULT OVERVIEW")

    ov_data = [
        [Paragraph("Total Students",    ov_key), Paragraph(str(len(df)), ov_val)],
        [Paragraph("Total Subjects",    ov_key), Paragraph(str(len(subj_cols)), ov_val)],
        [Paragraph("Average SGPA",      ov_key), Paragraph(str(avg_sgpa), ov_val)],
        [Paragraph("Highest SGPA",      ov_key), Paragraph(str(max_sgpa), ov_val)],
        [Paragraph("Lowest SGPA",       ov_key), Paragraph(str(min_sgpa), ov_val)],
        [Paragraph("Failures (SGPA 0)", ov_key),
         Paragraph(str(failures), ParagraphStyle(
             "F0", parent=ov_val, textColor=RED, alignment=TA_CENTER))],
    ]
    ov_tbl = Table(ov_data, colWidths=[AW * 0.46, AW * 0.2], hAlign="LEFT")
    ov_tbl.setStyle(TableStyle([
        ("GRID",          (0,0), (-1,-1), 0.5, MGRAY),
        ("ROWBACKGROUNDS",(0,0), (-1,-1), [rl.white, LGRAY]),
        ("TOPPADDING",    (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING",   (0,0), (0,-1),  8),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("LINEBELOW",     (0,-1),(-1,-1), 1.5, BLUE),
    ]))
    els.append(ov_tbl)
    els.append(Spacer(1, 6))

    # ══════════════════════════════════
    # § 3  COURSE DETAILS  (FIX #3 #8)
    # ══════════════════════════════════
    if course_map:
        els += section("COURSE DETAILS")
        cd_hdr = [Paragraph(h, th_s)
                  for h in ["S.No", "Subject Code", "Subject Name"]]
        cd_rows = []
        for i, (code, name) in enumerate(course_map, 1):
            cd_rows.append([
                Paragraph(str(i), td_sno),
                Paragraph(code,   cd_code),
                Paragraph(name,   cd_name),
            ])
        cd_tbl = Table(
            [cd_hdr] + cd_rows,
            colWidths=[24, AW * 0.20, AW * 0.60],
            hAlign="LEFT", repeatRows=1,
        )
        cd_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0),  NAVY),
            ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [rl.white, LGRAY]),
            ("TOPPADDING",    (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
            ("LINEBELOW",     (0,-1),(-1,-1), 1.5, BLUE),
        ]))
        els.append(cd_tbl)
        els.append(Spacer(1, 6))

    # ══════════════════════════════════
    # § 4  STUDENTS RESULT  (FIX #1 #7)
    # ══════════════════════════════════
    els += section("STUDENTS RESULT")

    # Fixed widths for anchor columns; subjects split what remains  (FIX #7)
    SNO_W  = 22
    ID_W   = 62 if id_col == "Register No" else 80
    SGPA_W = 30
    ARR_W  = 28
    remaining = AW - SNO_W - ID_W - SGPA_W - ARR_W
    SUBJ_W    = max(18, remaining / max(1, len(subj_cols)))

    col_ws = [SNO_W, ID_W] + [SUBJ_W] * len(subj_cols) + [SGPA_W, ARR_W]

    # ✅ Apply mode-based header label
    header_label = "Name" if mode == "Private" else "Register No"

    hdr_row = []
    for col in df.columns:
        if col == id_col:
            hdr_row.append(Paragraph(header_label, th_s))
        else:
            hdr_row.append(Paragraph(col, th_s))

    body_rows = []
    rs = []

    for i, (_, row) in enumerate(df.iterrows()):
        sgpa   = safe_float(row.get("SGPA", 0))
        ri     = i + 1
        row_bg = (_c("#FFEBEB") if sgpa == 0
                  else (_c("#FFFBEB") if 0 < sgpa < 5.0 else rl.white))
        rs.append(("BACKGROUND", (0, ri), (-1, ri), row_bg))

        cells = []
        for j, col in enumerate(df.columns):
            val = str(row[col]).strip()
            if val in ("nan", "None"):
                val = ""

            if col == "S.No":
                cells.append(Paragraph(val, td_sno))

            elif col == id_col:
                st2 = ParagraphStyle(f"R{i}", parent=td_l,
                                     fontName="Helvetica-Bold",
                                     textColor=RED if sgpa == 0 else DARK)
                cells.append(Paragraph(val, st2))

            elif col in subj_cols:
                bg_hex, fg_hex = _grade_colors(val.upper())
                if bg_hex and val:
                    rs.append(("BACKGROUND", (j, ri), (j, ri), _c(bg_hex)))
                    cells.append(Paragraph(val, ParagraphStyle(
                        f"G{i}{j}", parent=td_b, textColor=_c(fg_hex))))
                else:
                    cells.append(Paragraph(val, td_s))

            elif col == "SGPA":
                cells.append(Paragraph(val, ParagraphStyle(
                    f"S{i}", parent=td_b,
                    textColor=(RED if sgpa == 0
                               else (_c("#16A34A") if sgpa >= 7.5 else DARK)))))

            else:
                cells.append(Paragraph(val, td_s))

        body_rows.append(cells)

    base_style = [
        ("BACKGROUND",    (0,0), (-1,0),  NAVY),
        ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 2),
        ("RIGHTPADDING",  (0,0), (-1,-1), 2),
        ("REPEATROWS",    (0,0), (0,0)),
    ]
    stud_tbl = Table([hdr_row] + body_rows, colWidths=col_ws,
                     hAlign="LEFT", repeatRows=1)
    stud_tbl.setStyle(TableStyle(base_style + rs))
    els.append(stud_tbl)
    els.append(Spacer(1, 8))

    # ══════════════════════════════════
    # § 5  TOP PERFORMERS  (FIX #9)
    # ══════════════════════════════════
    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)

    top_hdr = [Paragraph(h, th_s) for h in ["Rank", id_col, "SGPA"]]
    top_rows = []
    for rank, (_, row) in enumerate(top_df.iterrows(), 1):
        top_rows.append([
            Paragraph(str(rank), top_rank),
            Paragraph(str(row.get(id_col, "")), td_l),
            Paragraph(str(row["SGPA"]), td_b),
        ])

    top_tbl = Table([top_hdr] + top_rows,
                    colWidths=[38, AW * 0.5, 48], hAlign="LEFT")
    top_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  _c("#D97706")),
        ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [rl.white, _c("#FFFBEB")]),
        ("LINEBELOW",     (0,-1),(-1,-1), 1.5, _c("#D97706")),
    ]))
    els.append(KeepTogether(section("TOP PERFORMERS") + [top_tbl, Spacer(1, 8)]))

    # ══════════════════════════════════
    # § 6  SUBJECT PERFORMANCE ANALYSIS
    # ══════════════════════════════════
    els += section("SUBJECT PERFORMANCE ANALYSIS")

    ana_cols = list(analysis_df.columns)
    SUBJ_NW  = 64
    OTHER_W  = (AW - SUBJ_NW) / max(1, len(ana_cols) - 1)
    ana_ws   = [SUBJ_NW] + [OTHER_W] * (len(ana_cols) - 1)

    ana_hdr  = [Paragraph(h, th_s) for h in ana_cols]
    ana_rows = []
    for _, row in analysis_df.iterrows():
        cells = []
        for col in ana_cols:
            v = str(row[col]) if str(row[col]) not in ("", "nan", "None") else "—"
            if col == "Subject":
                cells.append(Paragraph(v, td_l))
            elif "%" in v:
                try:
                    pct = float(v.replace("%", ""))
                    fg  = (RED if pct < 60
                           else (_c("#D97706") if pct < 80 else _c("#16A34A")))
                    cells.append(Paragraph(v, ParagraphStyle(
                        f"PP{col}", parent=td_b, textColor=fg)))
                except Exception:
                    cells.append(Paragraph(v, td_s))
            else:
                cells.append(Paragraph(v, td_s))
        ana_rows.append(cells)

    ana_tbl = Table([ana_hdr] + ana_rows, colWidths=ana_ws,
                    hAlign="LEFT", repeatRows=1)
    ana_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  NAVY),
        ("GRID",          (0,0), (-1,-1), 0.35, MGRAY),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 3),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [rl.white, LGRAY]),
        ("REPEATROWS",    (0,0), (0,0)),
        ("LINEBELOW",     (0,-1),(-1,-1), 1.5, BLUE),
    ]))
    els.append(ana_tbl)

    # ══════════════════════════════════
    # § 7  REPORT VERIFICATION
    # ══════════════════════════════════
    dept_s   = department if department else "the Department"
    ver_para = Paragraph(
        f"This is an officially verified result analysis document of the "
        f"{dept_s}, Carmel College of Engineering and Technology.", ver_body)

    if qr_path and os.path.exists(qr_path):
        ver_qr = Table(
            [[Image(qr_path, width=58, height=58)],
             [Paragraph("<para align='center'><font size='6.5' color='#64748B'>"
                        "Scan to verify</font></para>", S["Normal"])]],
            colWidths=[AW * 0.26],
        )
        ver_qr.setStyle(TableStyle([
            ("ALIGN",  (0,0), (-1,-1), "CENTER"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
    else:
        ver_qr = Paragraph("", S["Normal"])

    sig_tbl = Table(
        [[""], [Paragraph("Head of Department", sig_lbl)]],
        colWidths=[AW * 0.30], rowHeights=[30, None],
    )
    sig_tbl.setStyle(TableStyle([
        ("ALIGN",     (0,0), (-1,-1), "CENTER"),
        ("LINEABOVE", (0,1), (-1,1),  0.8, DARK),
        ("VALIGN",    (0,1), (0,1),   "TOP"),
    ]))

    ver_outer = Table(
        [[ver_para, ver_qr, sig_tbl]],
        colWidths=[AW * 0.44, AW * 0.26, AW * 0.30],
    )
    ver_outer.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
    ]))

    els.append(KeepTogether([
        Spacer(1, 22),
        HRFlowable(width="100%", thickness=1, color=BLUE,
                   spaceBefore=0, spaceAfter=6),
        Paragraph("Report Verification", ver_head),
        ver_outer,
        HRFlowable(width="100%", thickness=1, color=BLUE,
                   spaceBefore=4, spaceAfter=0),
    ]))

    doc.build(els, onFirstPage=_footer, onLaterPages=_footer)
    return output_path


# ═══════════════════════════════════════════════════════════════
# EXCEL REPORT   (FIX #1 #2 #3 #5 #8 #10 #11)
# ═══════════════════════════════════════════════════════════════

def generate_excel_report(
    df, analysis_df, output_path, report_id, qr_path,
    department="", semester="", batch="",
    db=None,# passed through; names fetched internally
    mode=None
):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # ── Shared preprocessing (FIX #4) ────────────────────────
    df, id_col, subj_cols = preprocess_df(df)
    course_map = build_course_map(subj_cols, db)   # FIX #3 #8

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Result Report"

    date_str = datetime.now().strftime("%d %B %Y")
    sem_lbl  = f"S{semester}" if semester else str(semester)

    THIN = Side(style="thin",   color="CBD5E1")

    def bdr():
        return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def cell(row, col, val, bold=False, size=9, color="0F172A",
             bg=None, align="left", va="center", wrap=False, border=True):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(horizontal=align, vertical=va, wrap_text=wrap)
        if bg:     c.fill   = PatternFill("solid", fgColor=bg.lstrip("#"))
        if border: c.border = bdr()
        return c

    # Total data columns after preprocessing
    TC = len(df.columns)
    r  = 1

    # ══════════════════════════════════
    # BLOCK 1 — HEADER
    # ══════════════════════════════════
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    c0 = ws.cell(row=r, column=1,
                 value="CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY")
    c0.font      = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    c0.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 22
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    c1 = ws.cell(row=r, column=1, value="RESULT ANALYSIS REPORT")
    c1.font      = Font(name="Calibri", size=11, bold=True, color="DC2626")
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 1

    # Separator
    for col in range(1, TC + 1):
        ws.cell(row=r, column=col).border = Border(
            bottom=Side(style="thin", color="CBD5E1"))
    ws.row_dimensions[r].height = 3
    r += 1

    # Meta block
    meta_pairs = [
        ("Report ID",  report_id),
        ("Date",       date_str),
        ("Department", department),
        ("Semester",   sem_lbl),
    ]
    if batch:
        meta_pairs.append(("Batch", batch))

    meta_start = r
    for key, val in meta_pairs:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ck = ws.cell(row=r, column=1, value=key)
        ck.font      = Font(name="Calibri", size=9, bold=True, color="0F172A")
        ck.alignment = Alignment(horizontal="right", vertical="center")
        cc = ws.cell(row=r, column=3, value=":")
        cc.font      = Font(name="Calibri", size=9, color="0F172A")
        cc.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=TC)
        cv = ws.cell(row=r, column=4, value=val)
        cv.font      = Font(name="Calibri", size=9, color="0F172A")
        cv.alignment = Alignment(horizontal="left", vertical="center",
                                 wrap_text=True)
        ws.row_dimensions[r].height = 14
        r += 1

    # Blue rule below meta
    for col in range(1, TC + 1):
        ex = ws.cell(row=r - 1, column=col).border
        ws.cell(row=r - 1, column=col).border = Border(
            left=ex.left, right=ex.right, top=ex.top,
            bottom=Side(style="medium", color="3B82F6"))
    r += 1
    ws.row_dimensions[r - 1].height = 4

    # QR image
    if qr_path and os.path.exists(qr_path):
        try:
            qr_img = XLImage(qr_path)
            qr_img.width = 65; qr_img.height = 65
            ws.add_image(qr_img, f"{get_column_letter(TC - 1)}1")
        except Exception:
            pass

    # ══════════════════════════════════
    # BLOCK 2 — RESULT OVERVIEW
    # ══════════════════════════════════
    sgpas    = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa = round(sum(sgpas) / len(sgpas), 2) if sgpas else 0
    max_sgpa = round(max(sgpas), 2) if sgpas else 0
    min_sgpa = round(min(sgpas), 2) if sgpas else 0
    failures = sum(1 for s in sgpas if s == 0)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl = ws.cell(row=r, column=1, value="RESULT OVERVIEW")
    sl.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl.alignment = Alignment(horizontal="left", vertical="center")
    sl.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    LBL_W, VAL_W = 5, 3
    for i, (lbl, val, vc) in enumerate([
        ("Total Students",    len(df),       "0F172A"),
        ("Total Subjects",    len(subj_cols), "0F172A"),
        ("Average SGPA",      avg_sgpa,       "0F172A"),
        ("Highest SGPA",      max_sgpa,       "16A34A"),
        ("Lowest SGPA",       min_sgpa,       "0F172A"),
        ("Failures (SGPA 0)", failures,       "DC2626"),
    ]):
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LBL_W)
        cl = ws.cell(row=r, column=1, value=lbl)
        cl.font      = Font(name="Calibri", size=9.5, bold=True, color="0F172A")
        cl.fill      = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cl.border    = bdr()
        ws.merge_cells(start_row=r, start_column=LBL_W+1,
                       end_row=r, end_column=LBL_W+VAL_W)
        cv2 = ws.cell(row=r, column=LBL_W+1, value=val)
        cv2.font      = Font(name="Calibri", size=9.5, bold=True, color=vc)
        cv2.fill      = PatternFill("solid", fgColor=bg)
        cv2.alignment = Alignment(horizontal="center", vertical="center")
        cv2.border    = bdr()
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1; ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════
    # BLOCK 3 — COURSE DETAILS  (FIX #3 #8)
    # ══════════════════════════════════
    if course_map:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
        sl_cd = ws.cell(row=r, column=1, value="COURSE DETAILS")
        sl_cd.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
        sl_cd.alignment = Alignment(horizontal="left", vertical="center")
        sl_cd.border    = Border(bottom=Side(style="thin", color="3B82F6"))
        ws.row_dimensions[r].height = 16
        r += 1

        # Header: S.No(col1) | Code(col2) | Name(col3…8)
        for ci, h in enumerate(["S.No", "Subject Code", "Subject Name"], 1):
            col_start = [1, 2, 3][ci-1]
            col_end   = [1, 2, min(8, TC)][ci-1]
            if col_start != col_end:
                ws.merge_cells(start_row=r, start_column=col_start,
                               end_row=r, end_column=col_end)
            c = ws.cell(row=r, column=col_start, value=h)
            c.font      = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
            c.fill      = PatternFill("solid", fgColor="1E3A8A")
            c.alignment = Alignment(horizontal="center" if ci < 3 else "left",
                                    vertical="center")
            c.border = bdr()
        ws.row_dimensions[r].height = 16
        r += 1

        for i, (code, name) in enumerate(course_map, 1):
            bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
            for ci, (col_s, col_e, val, al) in enumerate([
                (1, 1,            str(i),                     "center"),
                (2, 2,            code,                       "left"),
                (3, min(8, TC),   name,"left"),
            ], 1):
                if col_s != col_e:
                    ws.merge_cells(start_row=r, start_column=col_s,
                                   end_row=r, end_column=col_e)
                c = ws.cell(row=r, column=col_s, value=val)
                c.font      = Font(name="Calibri", size=8, bold=(ci == 2),
                                   color="0F172A")
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal=al, vertical="center",
                                        indent=1 if al == "left" else 0)
                c.border    = bdr()
            ws.row_dimensions[r].height = 14
            r += 1

        r += 1; ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════
    # BLOCK 4 — STUDENTS RESULT  (FIX #1 #5)
    # ══════════════════════════════════
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl2 = ws.cell(row=r, column=1, value="STUDENTS RESULT")
    sl2.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl2.alignment = Alignment(horizontal="left", vertical="center")
    sl2.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    header_label = "Name" if mode == "Private" else "Register No"

    for ci, col_name in enumerate(df.columns, 1):
        display_name = header_label if col_name == id_col else col_name
        c = ws.cell(row=r, column=ci, value=display_name)
        c = ws.cell(row=r, column=ci, value=col_name)
        c.font      = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = bdr()
    ws.row_dimensions[r].height = 18
    r += 1

    for _, row_data in df.iterrows():
        sgpa   = safe_float(row_data.get("SGPA", 0))
        row_bg = ("FFEBEB" if sgpa == 0
                  else ("FFFBEB" if 0 < sgpa < 5.0 else "FFFFFF"))

        for ci, col_name in enumerate(df.columns, 1):
            val = str(row_data[col_name]).strip()
            val = "" if val in ("nan", "None") else val

            if col_name == "S.No":
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, color="64748B")
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

            elif col_name == id_col:
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, bold=True,
                                   color="DC2626" if sgpa == 0 else "0F172A")
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        indent=1)
                c.border = bdr()

            elif col_name in subj_cols:
                bg_h, fg_h = _grade_colors(val.upper())
                c = ws.cell(row=r, column=ci, value=val if val else "")
                if bg_h and val:
                    c.fill = PatternFill("solid", fgColor=bg_h.lstrip("#"))
                    c.font = Font(name="Calibri", size=8, bold=True,
                                  color=fg_h.lstrip("#"))
                else:
                    c.fill = PatternFill("solid", fgColor=row_bg)
                    c.font = Font(name="Calibri", size=8, color="0F172A")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

            elif col_name == "SGPA":
                fg = ("DC2626" if sgpa == 0
                      else ("16A34A" if sgpa >= 7.5 else "0F172A"))
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, bold=True, color=fg)
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

            else:   # Arrear
                c = ws.cell(row=r, column=ci, value=val)
                c.font      = Font(name="Calibri", size=8, color="0F172A")
                c.fill      = PatternFill("solid", fgColor=row_bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = bdr()

        ws.row_dimensions[r].height = 13
        r += 1

    r += 1; ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════
    # BLOCK 5 — TOP PERFORMERS  (FIX #9)
    # ══════════════════════════════════
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl3 = ws.cell(row=r, column=1, value="TOP PERFORMERS")
    sl3.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl3.alignment = Alignment(horizontal="left", vertical="center")
    sl3.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    for ci, (h, cs, ce) in enumerate(
            [("Rank", 1, 2), (id_col, 3, 7), ("SGPA", 8, 10)], 1):
        ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
        c = ws.cell(row=r, column=cs, value=h)
        c.font      = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="D97706")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr()
    ws.row_dimensions[r].height = 16
    r += 1

    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)
    medals = {1: "D97706", 2: "6B7280", 3: "92400E"}

    for rank, (_, rd) in enumerate(top_df.iterrows(), 1):
        bg = "FFFBEB" if rank % 2 else "FFFFFF"
        m  = medals.get(rank, "0F172A")
        for cs, ce, val, al, color, bold in [
            (1, 2,  str(rank),              "center", m,        True),
            (3, 7,  str(rd.get(id_col,"")), "left",   "0F172A", False),
            (8, 10, str(rd["SGPA"]),        "center", "16A34A", True),
        ]:
            ws.merge_cells(start_row=r, start_column=cs, end_row=r, end_column=ce)
            c = ws.cell(row=r, column=cs, value=val)
            c.font      = Font(name="Calibri", size=9, bold=bold, color=color)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=al, vertical="center",
                                    indent=1 if al == "left" else 0)
            c.border    = bdr()
        ws.row_dimensions[r].height = 14
        r += 1

    r += 1; ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════
    # BLOCK 6 — SUBJECT PERFORMANCE ANALYSIS
    # ══════════════════════════════════
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl4 = ws.cell(row=r, column=1, value="SUBJECT PERFORMANCE ANALYSIS")
    sl4.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl4.alignment = Alignment(horizontal="left", vertical="center")
    sl4.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    ana_cols = list(analysis_df.columns)
    for ci, h in enumerate(ana_cols, 1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font      = Font(name="Calibri", size=8, bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="1E3A8A")
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                vertical="center", wrap_text=True,
                                indent=1 if ci == 1 else 0)
        c.border = bdr()
    ws.row_dimensions[r].height = 20
    r += 1

    for i, (_, rd) in enumerate(analysis_df.iterrows()):
        bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(ana_cols, 1):
            val = str(rd[col]) if str(rd[col]) not in ("", "nan", "None") else "—"
            c   = ws.cell(row=r, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(
                horizontal="left" if ci == 1 else "center",
                vertical="center", indent=1 if ci == 1 else 0)
            c.border = bdr()
            if ci == 1:
                c.font = Font(name="Calibri", size=8, bold=True, color="0F172A")
            elif "%" in val:
                try:
                    pct = float(val.replace("%", ""))
                    fg  = ("DC2626" if pct < 60
                           else ("D97706" if pct < 80 else "16A34A"))
                    c.font = Font(name="Calibri", size=8, bold=True, color=fg)
                except Exception:
                    c.font = Font(name="Calibri", size=8, color="0F172A")
            else:
                c.font = Font(name="Calibri", size=8, color="0F172A")
        ws.row_dimensions[r].height = 13
        r += 1

    r += 1; ws.row_dimensions[r - 1].height = 6

    # ══════════════════════════════════
    # BLOCK 7 — REPORT VERIFICATION
    # ══════════════════════════════════
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TC)
    sl5 = ws.cell(row=r, column=1, value="REPORT VERIFICATION")
    sl5.font      = Font(name="Calibri", size=10, bold=True, color="1E3A8A")
    sl5.alignment = Alignment(horizontal="left", vertical="center")
    sl5.border    = Border(bottom=Side(style="thin", color="3B82F6"))
    ws.row_dimensions[r].height = 16
    r += 1

    dept_s   = department if department else "the Department"
    ver_text = (f"This is an officially verified result analysis document "
                f"of the {dept_s}, Carmel College of Engineering "
                f"and Technology.")
    half = max(TC // 2, 5)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=half)
    cv3 = ws.cell(row=r, column=1, value=ver_text)
    cv3.font      = Font(name="Calibri", size=9, color="0F172A")
    cv3.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True, indent=1)
    cv3.border = bdr()

    sig_col = half + 2
    if sig_col <= TC:
        ws.merge_cells(start_row=r + 2, start_column=sig_col,
                       end_row=r + 2, end_column=TC)
        sg = ws.cell(row=r + 2, column=sig_col, value="Head of Department")
        sg.font      = Font(name="Calibri", size=9, bold=True, color="0F172A")
        sg.alignment = Alignment(horizontal="center", vertical="top")
        sg.border    = Border(top=Side(style="thin", color="0F172A"))
    for rr in range(r, r + 4):
        ws.row_dimensions[rr].height = 16

    # ══════════════════════════════════
    # COLUMN WIDTHS  (FIX #7)
    # ══════════════════════════════════
    ws.column_dimensions["A"].width = 5    # S.No
    ws.column_dimensions["B"].width = 16   # Register No / Student Name
    ws.column_dimensions["C"].width = 4    # meta colon / first subject col
    for ci in range(4, TC - 1):
        ws.column_dimensions[get_column_letter(ci)].width = 6.5
    if TC > 1:
        ws.column_dimensions[get_column_letter(TC - 1)].width = 5.5  # SGPA
        ws.column_dimensions[get_column_letter(TC)].width     = 5.5  # Arrear

    freeze_r = meta_start + len(meta_pairs) + 2
    ws.freeze_panes = f"B{freeze_r}"
    ws.page_setup.orientation              = "landscape"
    ws.page_setup.fitToWidth               = 1
    ws.page_setup.fitToHeight              = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"1:{meta_start + len(meta_pairs)}"

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════
# HTML REPORT   (FIX #1 #2 #3 #5 #6 #8 #10 #11)
# ═══════════════════════════════════════════════════════════════

def generate_html_report(
    df, analysis_df, report_id, department, semester, batch, qr_path,
    db=None,                    # passed through; names fetched internally
    mode=None
):
    # ── Shared preprocessing (FIX #4) ────────────────────────
    df, id_col, subj_cols = preprocess_df(df)
    course_map = build_course_map(subj_cols, db)   # FIX #3 #8

    date_str = datetime.now().strftime("%d %B %Y")
    sem_lbl  = f"S{semester}" if semester else str(semester)
    qr_uri   = _qr_b64(qr_path)

    sgpas    = [safe_float(x) for x in df["SGPA"]]
    avg_sgpa = round(sum(sgpas) / len(sgpas), 2) if sgpas else 0
    max_sgpa = round(max(sgpas), 2) if sgpas else 0
    min_sgpa = round(min(sgpas), 2) if sgpas else 0
    failures = sum(1 for s in sgpas if s == 0)

    # Top performers (FIX #9)
    top_df = df.copy()
    top_df["_n"] = top_df["SGPA"].apply(safe_float)
    top_df = top_df[top_df["_n"] > 0].sort_values("_n", ascending=False).head(3)

    # ── Grade cell ────────────────────────────────────────────
    def _gcell(grade):
        g    = str(grade).strip().upper()
        bg, fg = _grade_colors(g)
        disp = grade if grade not in ("", "nan", "None") else ""
        if bg and disp:
            return (f'<td style="background:{bg};color:{fg};font-weight:700;'
                    f'text-align:center;padding:3px 4px;white-space:nowrap;">'
                    f'{disp}</td>')
        return f'<td style="text-align:center;padding:3px 4px;">{disp or "&nbsp;"}</td>'

    # ── Student rows (FIX #6 — follows strict column order) ──
    student_rows = ""
    for _, row in df.iterrows():
        sgpa   = safe_float(row.get("SGPA", 0))
        row_bg = ("#FFEBEB" if sgpa == 0
                  else ("#FFFBEB" if 0 < sgpa < 5.0 else "#FFFFFF"))
        sno    = str(row.get("S.No", ""))
        ident  = str(row.get(id_col, ""))
        reg_c  = "#DC2626" if sgpa == 0 else "#0F172A"
        sgpa_c = ("#DC2626" if sgpa == 0
                  else ("#16A34A" if sgpa >= 7.5 else "#0F172A"))

        cells  = (f'<td style="background:{row_bg};color:#64748B;'
                  f'text-align:center;padding:3px 4px;font-size:11px;">{sno}</td>')
        cells += (f'<td style="background:{row_bg};color:{reg_c};'
                  f'font-weight:700;padding:4px 6px;white-space:nowrap;">{ident}</td>')
        for col in subj_cols:
            cells += _gcell(str(row.get(col, "")))
        cells += (f'<td style="background:{row_bg};color:{sgpa_c};'
                  f'font-weight:700;text-align:center;padding:4px;">'
                  f'{row.get("SGPA", "")}</td>')
        cells += (f'<td style="background:{row_bg};text-align:center;'
                  f'padding:4px;">{row.get("Arrear", "")}</td>')
        student_rows += f"<tr>{cells}</tr>\n"

    # ── Subject analysis rows ─────────────────────────────────
    ana_cols_list = list(analysis_df.columns)
    ana_rows = ""
    for i, (_, row) in enumerate(analysis_df.iterrows()):
        bg = "#F1F5F9" if i % 2 == 0 else "#FFFFFF"
        cells = ""
        for col in ana_cols_list:
            v     = str(row[col]) if str(row[col]) not in ("", "nan", "None") else "—"
            align = "left" if col == "Subject" else "center"
            style = f"background:{bg};padding:5px;text-align:{align};"
            if "%" in v:
                try:
                    pct = float(v.replace("%", ""))
                    fc  = ("#DC2626" if pct < 60
                           else ("#D97706" if pct < 80 else "#16A34A"))
                    style += f"color:{fc};font-weight:700;"
                except Exception:
                    pass
            cells += f'<td style="{style}">{v}</td>'
        ana_rows += f"<tr>{cells}</tr>\n"

    # ── Top performer rows ────────────────────────────────────
    top_rows = ""
    for rank, (_, row) in enumerate(top_df.iterrows(), 1):
        bg = "#FFFBEB" if rank % 2 else "#FFFFFF"
        top_rows += (
            f'<tr style="background:{bg};">'
            f'<td style="text-align:center;font-weight:700;color:#D97706;padding:6px;">{rank}</td>'
            f'<td style="padding:6px;">{row.get(id_col,"")}</td>'
            f'<td style="text-align:center;font-weight:700;padding:6px;">{row["SGPA"]}</td>'
            f'</tr>\n'
        )

    # ── Table headers (FIX #6 — fixed widths) ────────────────
    sno_th  = '<th style="width:36px;text-align:center;padding:5px;">S.No</th>'
    header_label = "Name" if mode == "Private" else "Register No"
    id_th = f'<th style="width:180px;text-align:left;padding:5px 6px;">{header_label}</th>'
    subj_th = "".join(
        f'<th style="min-width:42px;padding:5px 3px;white-space:nowrap;">{s}</th>'
        for s in subj_cols)
    tail_th = ('<th style="min-width:40px;padding:5px;">SGPA</th>'
               '<th style="min-width:36px;padding:5px;">Arrear</th>')
    ana_th  = "".join(f'<th style="padding:5px 3px;">{h}</th>'
                      for h in ana_cols_list)

    # ── Course Details block (FIX #3 #8) ─────────────────────
    course_details_html = ""
    if course_map:
        cd_rows = "".join(
            f'<tr style="background:{"#F1F5F9" if i % 2 == 0 else "#FFFFFF"};">'
            f'<td style="text-align:center;padding:5px 4px;color:#64748B;">{i}</td>'
            f'<td style="padding:5px 8px;font-weight:700;">{code}</td>'
            f'<td style="padding:5px 8px;">{name}</td>'
            f'</tr>'
            for i, (code, name) in enumerate(course_map, 1)
        )
        course_details_html = f"""
<div class="sect-title">COURSE DETAILS</div>
<table class="cd-tbl">
<thead><tr>
  <th style="width:40px;text-align:center;">S.No</th>
  <th style="width:120px;">Subject Code</th>
  <th>Subject Name</th>
</tr></thead>
<tbody>{cd_rows}</tbody>
</table>
"""

    qr_block = (f'<img src="{qr_uri}" style="width:80px;height:80px;" alt="QR"/>'
                if qr_uri else
                '<div style="width:80px;height:80px;border:1px dashed #cbd5e1;'
                'border-radius:4px;"></div>')

    dept_s    = department if department else "the Department"
    batch_row = (f"<tr><td class='meta-key'>Batch</td>"
                 f"<td class='meta-sep'>:</td><td>{batch}</td></tr>" if batch else "")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Result Analysis Report — {department}</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Segoe UI',Helvetica,Arial,sans-serif;
     color:#0F172A;background:#f8fafc;font-size:12px;}}
.wrap{{max-width:210mm;margin:0 auto;background:#fff;
       box-shadow:0 1px 8px rgba(0,0,0,.12);}}

/* Header */
.hdr-college{{background:#1E3A8A;color:#fff;text-align:center;
              padding:12px 16px;font-size:15px;font-weight:700;}}
.hdr-title{{background:#EFF6FF;text-align:center;padding:7px 16px;
            font-size:13px;font-weight:700;color:#DC2626;
            border-bottom:2.5px solid #3B82F6;}}
.hdr-meta{{display:flex;justify-content:space-between;align-items:flex-start;
           padding:12px 16px 10px;border-bottom:2px solid #3B82F6;}}
.meta-table{{border-collapse:collapse;font-size:9pt;}}
.meta-table td{{padding:2px 4px;vertical-align:top;}}
.meta-key{{font-weight:700;text-align:right;white-space:nowrap;}}
.meta-sep{{text-align:center;padding:2px 6px;}}
.qr-area{{text-align:center;}}
.qr-area p{{font-size:8px;color:#64748B;margin-top:4px;max-width:90px;}}

/* Body */
.body{{padding:0 16px 16px;}}
.sect-title{{font-size:11px;font-weight:700;color:#0F172A;
             border-bottom:1.5px solid #3B82F6;
             padding:10px 0 4px;margin-top:14px;margin-bottom:8px;}}

/* Overview */
.ov-table{{width:52%;border-collapse:collapse;}}
.ov-table td{{border:.5px solid #CBD5E1;padding:6px 10px;font-size:9.5pt;}}
.ov-table tr:nth-child(even){{background:#F1F5F9;}}
.ov-key{{font-weight:700;}}
.ov-val{{text-align:center;font-weight:700;}}

/* Course details */
.cd-tbl{{width:60%;border-collapse:collapse;font-size:9pt;margin-bottom:4px;}}
.cd-tbl th{{background:#1E3A8A;color:#fff;padding:5px 8px;text-align:left;}}
.cd-tbl td{{border:.4px solid #CBD5E1;}}
.cd-tbl tr:nth-child(even){{background:#F1F5F9;}}

/* Result table — FIX #6: table-layout:fixed prevents column drift */
.rtbl{{width:100%;border-collapse:collapse;font-size:8pt;table-layout:fixed;}}
.rtbl th{{background:#1E3A8A;color:#fff;padding:5px 3px;text-align:center;
          font-weight:600;white-space:nowrap;overflow:hidden;}}
.rtbl td{{border:.35px solid #CBD5E1;overflow:hidden;word-break:break-word;}}
.rtbl tr:nth-child(even) td:not([style*="background"]){{background:#F8FAFC;}}

/* Top performers */
.top-tbl{{width:55%;border-collapse:collapse;font-size:9pt;}}
.top-tbl th{{background:#D97706;color:#fff;padding:6px;text-align:center;}}
.top-tbl td{{border:.5px solid #CBD5E1;}}

/* Subject analysis */
.atbl{{width:100%;border-collapse:collapse;font-size:8pt;}}
.atbl th{{background:#1E3A8A;color:#fff;padding:5px 3px;text-align:center;}}
.atbl tr:nth-child(even){{background:#F1F5F9;}}
.atbl td{{border:.35px solid #CBD5E1;padding:4px 3px;text-align:center;}}
.atbl td:first-child{{text-align:left;padding-left:6px;}}

/* Verification */
.ver{{display:flex;align-items:center;gap:0;border:1px solid #E2E8F0;
      border-radius:6px;padding:12px;margin-top:16px;}}
.ver-text{{flex:1;font-size:9pt;line-height:1.5;}}
.ver-title{{font-weight:700;font-size:11pt;color:#1D4ED8;margin-bottom:6px;}}
.ver-qr{{text-align:center;padding:0 20px;}}
.ver-qr p{{font-size:7.5pt;color:#64748B;margin-top:4px;}}
.ver-sig{{flex:1;text-align:center;}}
.ver-sig .line{{border-top:1px solid #0F172A;width:75%;margin:28px auto 6px;}}
.ver-sig .lbl{{font-weight:700;font-size:9pt;}}

@media print{{
  body{{background:#fff;-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
  .wrap{{box-shadow:none;max-width:100%;}}
  .page-break{{page-break-before:always;}}
  .rtbl{{font-size:7pt;}}
}}
</style>
</head>
<body>
<div class="wrap">

<div class="hdr-college">CARMEL COLLEGE OF ENGINEERING AND TECHNOLOGY</div>
<div class="hdr-title">RESULT ANALYSIS REPORT</div>

<div class="hdr-meta">
  <table class="meta-table">
    <tr><td class="meta-key">Report ID</td><td class="meta-sep">:</td><td>{report_id}</td></tr>
    <tr><td class="meta-key">Date</td><td class="meta-sep">:</td><td>{date_str}</td></tr>
    <tr><td class="meta-key">Department</td><td class="meta-sep">:</td><td>{department}</td></tr>
    <tr><td class="meta-key">Semester</td><td class="meta-sep">:</td><td>{sem_lbl}</td></tr>
    {batch_row}
  </table>
  <div class="qr-area">{qr_block}<p>Scan to verify<br>authenticity</p></div>
</div>

<div class="body">

<div class="sect-title">RESULT OVERVIEW</div>
<table class="ov-table"><tbody>
  <tr><td class="ov-key">Total Students</td><td class="ov-val">{len(df)}</td></tr>
  <tr><td class="ov-key">Total Subjects</td><td class="ov-val">{len(subj_cols)}</td></tr>
  <tr><td class="ov-key">Average SGPA</td><td class="ov-val">{avg_sgpa}</td></tr>
  <tr><td class="ov-key">Highest SGPA</td><td class="ov-val" style="color:#16A34A;">{max_sgpa}</td></tr>
  <tr><td class="ov-key">Lowest SGPA</td><td class="ov-val">{min_sgpa}</td></tr>
  <tr><td class="ov-key">Failures (SGPA 0)</td><td class="ov-val" style="color:#DC2626;">{failures}</td></tr>
</tbody></table>

{course_details_html}

<div class="sect-title">STUDENTS RESULT</div>
<table class="rtbl">
<thead><tr>{sno_th}{id_th}{subj_th}{tail_th}</tr></thead>
<tbody>{student_rows}</tbody>
</table>

<div class="sect-title">TOP PERFORMERS</div>
<table class="top-tbl">
<thead><tr><th>Rank</th><th>{id_col}</th><th>SGPA</th></tr></thead>
<tbody>{top_rows}</tbody>
</table>

<div class="sect-title page-break" style="margin-top:20px;">SUBJECT PERFORMANCE ANALYSIS</div>
<table class="atbl">
<thead><tr>{ana_th}</tr></thead>
<tbody>{ana_rows}</tbody>
</table>

<div class="ver">
  <div class="ver-text">
    <div class="ver-title">Report Verification</div>
    This is an officially verified result analysis document of the
    {dept_s}, Carmel College of Engineering and Technology.
  </div>
  <div class="ver-qr">{qr_block}<p>Scan to verify</p></div>
  <div class="ver-sig"><div class="line"></div><div class="lbl">Head of Department</div></div>
</div>

</div>
</div>
</body>
</html>"""

    return html
